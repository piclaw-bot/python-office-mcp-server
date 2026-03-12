"""
Tests for excel_advanced_tools.py - Advanced Excel workbook manipulation

Tests cover:
- Workbook introspection (list_sheets, list_named_ranges, list_tables)
- Cell/range operations (get_range, patch_cell, patch_range)
- Table operations (get_table, append_table_row, update_table_row)
- Placeholder operations (replace_placeholders, audit_placeholders)
- Template and change log operations
"""

import pytest

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from tools.excel_advanced_tools import (
    _add_change_comment,
    _ensure_change_log_sheet,
    _get_range_bounds,
    _highlight_cell,
    _log_change,
    _parse_cell_reference,
)

# Fixtures temp_dir and excel_advanced_tools are provided by conftest.py


@pytest.fixture
def simple_xlsx(temp_dir):
    """Create a simple Excel workbook for testing."""
    if not HAS_OPENPYXL:
        pytest.skip("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Add some data
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "Status"
    ws["A2"] = "Item 1"
    ws["B2"] = 100
    ws["C2"] = "Active"
    ws["A3"] = "Item 2"
    ws["B3"] = 200
    ws["C3"] = "Inactive"

    # Add a second sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Total"
    ws2["B1"] = "=SUM(Data!B:B)"

    path = temp_dir / "simple.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_xlsx(temp_dir):
    """Create workbook with multiple sheets including hidden ones."""
    if not HAS_OPENPYXL:
        pytest.skip("openpyxl not installed")

    wb = Workbook()

    # Visible sheet
    ws1 = wb.active
    ws1.title = "Main"
    ws1["A1"] = "Main content"

    # Hidden sheet
    ws2 = wb.create_sheet("Hidden")
    ws2.sheet_state = "hidden"
    ws2["A1"] = "Hidden content"

    # Very hidden sheet
    ws3 = wb.create_sheet("VeryHidden")
    ws3.sheet_state = "veryHidden"
    ws3["A1"] = "Secret content"

    path = temp_dir / "multi_sheet.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_with_placeholders(temp_dir):
    """Create workbook with various placeholder patterns."""
    if not HAS_OPENPYXL:
        pytest.skip("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Form"

    ws["A1"] = "Customer Name:"
    ws["B1"] = "<Customer Name>"
    ws["A2"] = "Project:"
    ws["B2"] = "<Project Name>"
    ws["A3"] = "Status:"
    ws["B3"] = "[TBD]"
    ws["A4"] = "Date:"
    ws["B4"] = "[enter date]"
    ws["A5"] = "Notes:"
    ws["B5"] = "Contact: <Customer Name> about <Project Name>"

    path = temp_dir / "placeholders.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_with_table(temp_dir):
    """Create workbook with an Excel table."""
    if not HAS_OPENPYXL:
        pytest.skip("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "TaskList"

    # Table headers
    ws["A1"] = "Task"
    ws["B1"] = "Owner"
    ws["C1"] = "Status"
    ws["D1"] = "Due Date"

    # Table data
    ws["A2"] = "Design"
    ws["B2"] = "Alice"
    ws["C2"] = "Complete"
    ws["D2"] = "2026-01-15"

    ws["A3"] = "Build"
    ws["B3"] = "Bob"
    ws["C3"] = "In Progress"
    ws["D3"] = "2026-02-28"

    ws["A4"] = "Test"
    ws["B4"] = "Carol"
    ws["C4"] = "Pending"
    ws["D4"] = "2026-03-15"

    # Create table
    table = Table(displayName="Tasks", ref="A1:D4")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    path = temp_dir / "with_table.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_with_named_ranges(temp_dir):
    """Create workbook with named ranges."""
    if not HAS_OPENPYXL:
        pytest.skip("openpyxl not installed")

    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = "Customer"
    ws["A2"] = "Contoso"
    ws["B1"] = "Amount"
    ws["B2"] = 10000

    # Workbook-scoped named ranges
    wb.defined_names.add(DefinedName("CustomerName", attr_text="Data!$A$2"))
    wb.defined_names.add(DefinedName("Amount", attr_text="Data!$B$2"))
    wb.defined_names.add(DefinedName("DataRange", attr_text="Data!$A$1:$B$2"))

    path = temp_dir / "named_ranges.xlsx"
    wb.save(path)
    return path


# =============================================================================
# HELPER FUNCTION TESTS
# =============================================================================


class TestHelperFunctions:
    """Tests for module-level helper functions."""

    def test_parse_cell_reference_simple(self):
        """Should parse simple cell reference."""
        sheet, cell = _parse_cell_reference("A1")
        assert sheet is None
        assert cell == "A1"

    def test_parse_cell_reference_with_sheet(self):
        """Should parse cell reference with sheet name."""
        sheet, cell = _parse_cell_reference("Sheet1!B5")
        assert sheet == "Sheet1"
        assert cell == "B5"

    def test_parse_cell_reference_quoted_sheet(self):
        """Should handle quoted sheet names."""
        sheet, cell = _parse_cell_reference("'My Sheet'!C10")
        assert sheet == "My Sheet"
        assert cell == "C10"

    def test_get_range_bounds_single_cell(self):
        """Should parse single cell as 1x1 range."""
        min_row, min_col, max_row, max_col = _get_range_bounds("B5")
        assert min_row == 5
        assert min_col == 2
        assert max_row == 5
        assert max_col == 2

    def test_get_range_bounds_range(self):
        """Should parse rectangular range."""
        min_row, min_col, max_row, max_col = _get_range_bounds("A1:D10")
        assert min_row == 1
        assert min_col == 1
        assert max_row == 10
        assert max_col == 4

    def test_get_range_bounds_invalid(self):
        """Should raise error for invalid reference."""
        with pytest.raises(ValueError):
            _get_range_bounds("invalid")

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_ensure_change_log_sheet(self):
        """Should create _ChangeLog sheet if missing."""
        wb = Workbook()
        assert "_ChangeLog" not in wb.sheetnames

        _ensure_change_log_sheet(wb)
        assert "_ChangeLog" in wb.sheetnames

        # Check headers
        ws = wb["_ChangeLog"]
        assert ws["A1"].value == "Timestamp"
        assert ws["F1"].value == "Author"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_log_change(self):
        """Should log change to _ChangeLog sheet."""
        wb = Workbook()
        _log_change(wb, "Sheet1", "A1", "old", "new", "TestAuthor")

        assert "_ChangeLog" in wb.sheetnames
        ws = wb["_ChangeLog"]
        assert ws["B2"].value == "Sheet1"
        assert ws["C2"].value == "A1"
        assert ws["D2"].value == "old"
        assert ws["E2"].value == "new"
        assert ws["F2"].value == "TestAuthor"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_highlight_cell(self):
        """Should apply yellow highlight to cell."""
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.value = "Test"

        _highlight_cell(cell)

        # openpyxl may add alpha prefix (00) to RGB
        assert cell.fill.start_color.rgb.endswith("FFFF99")

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_add_change_comment(self):
        """Should add comment documenting change."""
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.value = "Test"

        _add_change_comment(cell, "old_value", "new_value", "TestAuthor")

        assert cell.comment is not None
        assert "TestAuthor" in cell.comment.text
        assert "old_value" in cell.comment.text
        assert "new_value" in cell.comment.text


# =============================================================================
# WORKBOOK INTROSPECTION TESTS
# =============================================================================


class TestListSheets:
    """Tests for tool_excel_list_sheets."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_lists_all_sheets(self, excel_advanced_tools, multi_sheet_xlsx):
        """Should list all sheets including hidden ones."""
        result = excel_advanced_tools.tool_excel_list_sheets(str(multi_sheet_xlsx))
        assert result["sheet_count"] == 3
        names = [s["name"] for s in result["sheets"]]
        assert "Main" in names
        assert "Hidden" in names
        assert "VeryHidden" in names

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_excludes_hidden_when_requested(self, excel_advanced_tools, multi_sheet_xlsx):
        """Should exclude hidden sheets when include_hidden=False."""
        result = excel_advanced_tools.tool_excel_list_sheets(
            str(multi_sheet_xlsx), include_hidden=False
        )
        assert result["sheet_count"] == 1
        assert result["sheets"][0]["name"] == "Main"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_includes_sheet_state(self, excel_advanced_tools, multi_sheet_xlsx):
        """Should include visibility state for each sheet."""
        result = excel_advanced_tools.tool_excel_list_sheets(str(multi_sheet_xlsx))
        states = {s["name"]: s["state"] for s in result["sheets"]}
        assert states["Main"] == "visible"
        assert states["Hidden"] == "hidden"
        assert states["VeryHidden"] == "veryHidden"

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_list_sheets("/nonexistent.xlsx")
        assert "error" in result


class TestListNamedRanges:
    """Tests for tool_excel_list_named_ranges."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_lists_named_ranges(self, excel_advanced_tools, xlsx_with_named_ranges):
        """Should list all named ranges."""
        result = excel_advanced_tools.tool_excel_list_named_ranges(str(xlsx_with_named_ranges))
        assert result["count"] >= 2
        names = [nr["name"] for nr in result["named_ranges"]]
        assert "CustomerName" in names
        assert "Amount" in names

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_includes_scope(self, excel_advanced_tools, xlsx_with_named_ranges):
        """Should identify workbook scope."""
        result = excel_advanced_tools.tool_excel_list_named_ranges(str(xlsx_with_named_ranges))
        scopes = {nr["name"]: nr["scope"] for nr in result["named_ranges"]}
        assert scopes["CustomerName"] == "Workbook"
        assert scopes["DataRange"] == "Workbook"

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_list_named_ranges("/nonexistent.xlsx")
        assert "error" in result


class TestListTables:
    """Tests for tool_excel_list_tables."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_lists_tables(self, excel_advanced_tools, xlsx_with_table):
        """Should list all tables."""
        result = excel_advanced_tools.tool_excel_list_tables(str(xlsx_with_table))
        assert result["table_count"] == 1
        assert result["tables"][0]["name"] == "Tasks"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_includes_columns(self, excel_advanced_tools, xlsx_with_table):
        """Should include column names."""
        result = excel_advanced_tools.tool_excel_list_tables(str(xlsx_with_table))
        cols = result["tables"][0]["columns"]
        assert "Task" in cols
        assert "Owner" in cols
        assert "Status" in cols

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_filter_by_sheet(self, excel_advanced_tools, xlsx_with_table):
        """Should filter tables by sheet name."""
        result = excel_advanced_tools.tool_excel_list_tables(
            str(xlsx_with_table), sheet_name="NonExistent"
        )
        assert result["table_count"] == 0

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_list_tables("/nonexistent.xlsx")
        assert "error" in result


# =============================================================================
# CELL/RANGE OPERATION TESTS
# =============================================================================


class TestGetRange:
    """Tests for tool_excel_get_range."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_reads_single_cell(self, excel_advanced_tools, simple_xlsx):
        """Should read a single cell."""
        result = excel_advanced_tools.tool_excel_get_range(str(simple_xlsx), "A1", sheet_name="Data")
        assert result["row_count"] == 1
        assert result["col_count"] == 1
        assert result["data"][0][0]["value"] == "Name"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_reads_range(self, excel_advanced_tools, simple_xlsx):
        """Should read rectangular range."""
        result = excel_advanced_tools.tool_excel_get_range(
            str(simple_xlsx), "A1:C2", sheet_name="Data"
        )
        assert result["row_count"] == 2
        assert result["col_count"] == 3

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_includes_formulas(self, excel_advanced_tools, simple_xlsx):
        """Should include formulas when requested."""
        result = excel_advanced_tools.tool_excel_get_range(
            str(simple_xlsx), "B1", sheet_name="Summary", include_formulas=True
        )
        # The formula should be preserved
        value = result["data"][0][0]["value"]
        assert value is not None

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_sheet_in_reference(self, excel_advanced_tools, simple_xlsx):
        """Should parse sheet name from reference."""
        result = excel_advanced_tools.tool_excel_get_range(str(simple_xlsx), "Data!A1")
        assert result["sheet"] == "Data"

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_get_range("/nonexistent.xlsx", "A1")
        assert "error" in result

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_sheet_not_found(self, excel_advanced_tools, simple_xlsx):
        """Should handle missing sheet."""
        result = excel_advanced_tools.tool_excel_get_range(
            str(simple_xlsx), "A1", sheet_name="NonExistent"
        )
        assert "error" in result


class TestPatchCell:
    """Tests for tool_excel_patch_cell."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_updates_cell(self, excel_advanced_tools, simple_xlsx):
        """Should update cell value."""
        result = excel_advanced_tools.tool_excel_patch_cell(
            str(simple_xlsx), "A2", "Updated", sheet_name="Data"
        )
        assert result["success"] is True
        assert result["old_value"] == "Item 1"
        assert result["new_value"] == "Updated"

        # Verify change persisted
        from openpyxl import load_workbook

        wb = load_workbook(simple_xlsx)
        assert wb["Data"]["A2"].value == "Updated"
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_creates_change_log(self, excel_advanced_tools, simple_xlsx):
        """Should log change to _ChangeLog sheet."""
        excel_advanced_tools.tool_excel_patch_cell(
            str(simple_xlsx), "A2", "Updated", sheet_name="Data"
        )

        from openpyxl import load_workbook

        wb = load_workbook(simple_xlsx)
        assert "_ChangeLog" in wb.sheetnames
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_highlights_cell(self, excel_advanced_tools, simple_xlsx):
        """Should highlight changed cell."""
        excel_advanced_tools.tool_excel_patch_cell(
            str(simple_xlsx), "A2", "Updated", sheet_name="Data", highlight=True
        )

        from openpyxl import load_workbook

        wb = load_workbook(simple_xlsx)
        cell = wb["Data"]["A2"]
        # openpyxl may add alpha prefix (00) to RGB
        assert cell.fill.start_color.rgb.endswith("FFFF99")
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_respects_output_path(self, excel_advanced_tools, simple_xlsx, temp_dir):
        """Should save to output_path when specified."""
        output = temp_dir / "patched.xlsx"
        result = excel_advanced_tools.tool_excel_patch_cell(
            str(simple_xlsx), "A2", "Updated", sheet_name="Data", output_path=str(output)
        )
        assert result["file"] == str(output)
        assert output.exists()

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_patch_cell("/nonexistent.xlsx", "A1", "value")
        assert "error" in result


class TestPatchRange:
    """Tests for tool_excel_patch_range."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_updates_range(self, excel_advanced_tools, simple_xlsx):
        """Should update multiple cells."""
        values = [["New1", 111, "Status1"], ["New2", 222, "Status2"]]
        result = excel_advanced_tools.tool_excel_patch_range(
            str(simple_xlsx), "A2:C3", values, sheet_name="Data"
        )
        assert result["success"] is True
        assert result["cells_changed"] > 0

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_validates_dimensions(self, excel_advanced_tools, simple_xlsx):
        """Should reject mismatched dimensions."""
        values = [["Only", "Two"]]  # 1 row instead of expected 2
        result = excel_advanced_tools.tool_excel_patch_range(
            str(simple_xlsx), "A2:C3", values, sheet_name="Data"
        )
        assert "error" in result

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_logs_all_changes(self, excel_advanced_tools, simple_xlsx):
        """Should log each cell change."""
        values = [["New1", 111, "Status1"]]
        excel_advanced_tools.tool_excel_patch_range(
            str(simple_xlsx), "A2:C2", values, sheet_name="Data"
        )

        from openpyxl import load_workbook

        wb = load_workbook(simple_xlsx)
        ws = wb["_ChangeLog"]
        # Count non-empty rows (excluding header)
        log_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
        assert len(log_rows) >= 1
        wb.close()


# =============================================================================
# TABLE OPERATION TESTS
# =============================================================================


class TestGetTable:
    """Tests for tool_excel_get_table."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_reads_table_data(self, excel_advanced_tools, xlsx_with_table):
        """Should read all table data."""
        result = excel_advanced_tools.tool_excel_get_table(str(xlsx_with_table), "Tasks")
        assert result["table_name"] == "Tasks"
        assert result["row_count"] >= 3  # Header + 3 data rows

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_includes_columns(self, excel_advanced_tools, xlsx_with_table):
        """Should include column names."""
        result = excel_advanced_tools.tool_excel_get_table(str(xlsx_with_table), "Tasks")
        assert "Task" in result["columns"]
        assert "Owner" in result["columns"]

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_excludes_headers(self, excel_advanced_tools, xlsx_with_table):
        """Should exclude headers when requested."""
        result = excel_advanced_tools.tool_excel_get_table(
            str(xlsx_with_table), "Tasks", include_headers=False
        )
        # First row should be data, not header
        assert result["data"][0][0] != "Task"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_table_not_found(self, excel_advanced_tools, xlsx_with_table):
        """Should handle missing table."""
        result = excel_advanced_tools.tool_excel_get_table(str(xlsx_with_table), "NonExistent")
        assert "error" in result


class TestAppendTableRow:
    """Tests for tool_excel_append_table_row."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_appends_row(self, excel_advanced_tools, xlsx_with_table):
        """Should append new row to table."""
        result = excel_advanced_tools.tool_excel_append_table_row(
            str(xlsx_with_table),
            "Tasks",
            {"Task": "Deploy", "Owner": "Dave", "Status": "Pending", "Due Date": "2026-04-01"},
        )
        assert result["success"] is True
        assert result["new_row"] == 5  # Was 4 rows, now 5

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_expands_table_range(self, excel_advanced_tools, xlsx_with_table):
        """Should expand table range."""
        result = excel_advanced_tools.tool_excel_append_table_row(
            str(xlsx_with_table), "Tasks", {"Task": "New Task"}
        )
        assert "D5" in result["new_range"]

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_handles_partial_data(self, excel_advanced_tools, xlsx_with_table):
        """Should handle missing columns."""
        result = excel_advanced_tools.tool_excel_append_table_row(
            str(xlsx_with_table), "Tasks", {"Task": "Partial"}
        )
        assert result["success"] is True
        assert "Task" in result["columns_filled"]


class TestUpdateTableRow:
    """Tests for tool_excel_update_table_row."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_updates_row(self, excel_advanced_tools, xlsx_with_table):
        """Should update specific row."""
        result = excel_advanced_tools.tool_excel_update_table_row(
            str(xlsx_with_table), "Tasks", 1, {"Status": "Done"}
        )
        assert result["success"] is True
        assert len(result["updates"]) == 1
        assert result["updates"][0]["column"] == "Status"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_row_out_of_range(self, excel_advanced_tools, xlsx_with_table):
        """Should handle row index out of range."""
        result = excel_advanced_tools.tool_excel_update_table_row(
            str(xlsx_with_table), "Tasks", 100, {"Status": "Done"}
        )
        assert "error" in result

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_logs_and_highlights(self, excel_advanced_tools, xlsx_with_table):
        """Should log and highlight changes."""
        excel_advanced_tools.tool_excel_update_table_row(
            str(xlsx_with_table), "Tasks", 1, {"Status": "Done"}
        )

        from openpyxl import load_workbook

        wb = load_workbook(xlsx_with_table)
        assert "_ChangeLog" in wb.sheetnames
        wb.close()


# =============================================================================
# PLACEHOLDER OPERATION TESTS
# =============================================================================


class TestReplacePlaceholders:
    """Tests for tool_excel_replace_placeholders."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_replaces_placeholders(self, excel_advanced_tools, xlsx_with_placeholders):
        """Should replace placeholder patterns."""
        result = excel_advanced_tools.tool_excel_replace_placeholders(
            str(xlsx_with_placeholders),
            {"<Customer Name>": "Contoso", "<Project Name>": "Migration"},
        )
        assert result["success"] is True
        assert result["total_replacements"] >= 2

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_counts_by_placeholder(self, excel_advanced_tools, xlsx_with_placeholders):
        """Should count replacements per placeholder."""
        result = excel_advanced_tools.tool_excel_replace_placeholders(
            str(xlsx_with_placeholders),
            {"<Customer Name>": "Contoso"},
        )
        # <Customer Name> appears in B1 and B5 (twice in B5)
        assert result["by_placeholder"]["<Customer Name>"] >= 2

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_handles_partial_replacement(self, excel_advanced_tools, xlsx_with_placeholders):
        """Should replace placeholder within text."""
        excel_advanced_tools.tool_excel_replace_placeholders(
            str(xlsx_with_placeholders),
            {"<Customer Name>": "Contoso", "<Project Name>": "Migration"},
        )

        from openpyxl import load_workbook

        wb = load_workbook(xlsx_with_placeholders)
        # B5 had "Contact: <Customer Name> about <Project Name>"
        assert "Contoso" in wb["Form"]["B5"].value
        assert "Migration" in wb["Form"]["B5"].value
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_filters_by_sheet(self, excel_advanced_tools, xlsx_with_placeholders):
        """Should limit to specified sheets."""
        result = excel_advanced_tools.tool_excel_replace_placeholders(
            str(xlsx_with_placeholders),
            {"<Customer Name>": "Contoso"},
            sheet_names=["NonExistent"],
        )
        assert result["total_replacements"] == 0


class TestAuditPlaceholders:
    """Tests for tool_excel_audit_placeholders."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_finds_placeholders(self, excel_advanced_tools, xlsx_with_placeholders):
        """Should find placeholder patterns."""
        result = excel_advanced_tools.tool_excel_audit_placeholders(str(xlsx_with_placeholders))
        assert result["total_found"] > 0
        assert result["status"] == "needs_attention"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_reports_locations(self, excel_advanced_tools, xlsx_with_placeholders):
        """Should report cell locations."""
        result = excel_advanced_tools.tool_excel_audit_placeholders(str(xlsx_with_placeholders))
        cells = [f["cell"] for f in result["findings"]]
        assert "B1" in cells or "B2" in cells  # Where placeholders are

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_custom_patterns(self, excel_advanced_tools, xlsx_with_placeholders):
        """Should search for custom patterns."""
        result = excel_advanced_tools.tool_excel_audit_placeholders(
            str(xlsx_with_placeholders), patterns=["[TBD]"]
        )
        assert result["total_found"] >= 1

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_clean_workbook(self, excel_advanced_tools, simple_xlsx):
        """Should report clean status for workbook without placeholders."""
        result = excel_advanced_tools.tool_excel_audit_placeholders(str(simple_xlsx))
        assert result["status"] == "clean"


# =============================================================================
# TEMPLATE AND CHANGE LOG TESTS
# =============================================================================


class TestCopyTemplate:
    """Tests for tool_excel_copy_template."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_copies_file(self, excel_advanced_tools, simple_xlsx, temp_dir):
        """Should copy template to new location."""
        output = temp_dir / "copy.xlsx"
        result = excel_advanced_tools.tool_excel_copy_template(str(simple_xlsx), str(output))
        assert result["success"] is True
        assert output.exists()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_preserves_content(self, excel_advanced_tools, simple_xlsx, temp_dir):
        """Should preserve all content."""
        output = temp_dir / "copy.xlsx"
        excel_advanced_tools.tool_excel_copy_template(str(simple_xlsx), str(output))

        from openpyxl import load_workbook

        wb = load_workbook(output)
        assert wb["Data"]["A1"].value == "Name"
        wb.close()

    def test_template_not_found(self, excel_advanced_tools, temp_dir):
        """Should handle missing template."""
        result = excel_advanced_tools.tool_excel_copy_template(
            "/nonexistent.xlsx", str(temp_dir / "out.xlsx")
        )
        assert "error" in result


class TestGetChangeLog:
    """Tests for tool_excel_get_change_log."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_no_change_log(self, excel_advanced_tools, simple_xlsx):
        """Should handle workbook without change log."""
        result = excel_advanced_tools.tool_excel_get_change_log(str(simple_xlsx))
        assert result["has_change_log"] is False

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_retrieves_log(self, excel_advanced_tools, simple_xlsx):
        """Should retrieve change log entries."""
        # First make a change to create the log
        excel_advanced_tools.tool_excel_patch_cell(
            str(simple_xlsx), "A2", "Changed", sheet_name="Data"
        )

        result = excel_advanced_tools.tool_excel_get_change_log(str(simple_xlsx))
        assert result["has_change_log"] is True
        assert result["entry_count"] >= 1
        assert result["entries"][0]["new_value"] == "Changed"

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_get_change_log("/nonexistent.xlsx")
        assert "error" in result


# =============================================================================
# REAL-WORLD TEMPLATE TESTS
# =============================================================================


class TestECIFTemplate:
    """Tests using an ECIF-like generated XLSM template."""

    @pytest.fixture
    def ecif_copy(self, temp_dir):
        """Create a local ECIF-like workbook for testing."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl not installed")

        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()

        ws = wb.active
        ws.title = "ECIF Work Scope (E)"
        ws["A1"] = "Milestone #"
        ws["B1"] = "Description"
        ws["C1"] = "Owner"
        ws["A2"] = "M1"
        ws["B2"] = "Initiate"
        ws["C2"] = "Alice"
        ws["A3"] = "M2"
        ws["B3"] = "Design"
        ws["C3"] = "Bob"
        ws["A4"] = "M3"
        ws["B4"] = "Deploy"
        ws["C4"] = "Carol"

        table = Table(displayName="Milestones", ref="A1:C4")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        for sheet_name in [
            "Read Me",
            "Scope",
            "Assumptions",
            "Pricing",
            "Appendix",
            "Reference",
        ]:
            wb.create_sheet(sheet_name)

        wb.defined_names.add(DefinedName("CustomerName", attr_text="'ECIF Work Scope (E)'!$B$5"))
        wb.defined_names.add(DefinedName("ProjectName", attr_text="'ECIF Work Scope (E)'!$B$6"))
        wb.defined_names.add(DefinedName("WorkOrder", attr_text="'ECIF Work Scope (E)'!$B$7"))
        wb.defined_names.add(DefinedName("DeliveryLead", attr_text="'ECIF Work Scope (E)'!$B$8"))
        wb.defined_names.add(DefinedName("Country", attr_text="'ECIF Work Scope (E)'!$B$9"))
        wb.defined_names.add(DefinedName("Region", attr_text="'ECIF Work Scope (E)'!$B$10"))
        wb.defined_names.add(DefinedName("StartDate", attr_text="'ECIF Work Scope (E)'!$B$11"))
        wb.defined_names.add(DefinedName("EndDate", attr_text="'ECIF Work Scope (E)'!$B$12"))
        wb.defined_names.add(DefinedName("TotalHours", attr_text="'ECIF Work Scope (E)'!$B$13"))
        wb.defined_names.add(DefinedName("BudgetUSD", attr_text="'ECIF Work Scope (E)'!$B$14"))
        wb.defined_names.add(DefinedName("MilestoneTable", attr_text="'ECIF Work Scope (E)'!$A$1:$C$4"))
        wb.defined_names.add(DefinedName("StatusList", attr_text="'ECIF Work Scope (E)'!$D$2:$D$4"))

        dest = temp_dir / "ecif_test.xlsm"
        wb.save(dest)
        return dest

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_list_ecif_sheets(self, excel_advanced_tools, ecif_copy):
        """Should list ECIF sheets including hidden ones."""
        result = excel_advanced_tools.tool_excel_list_sheets(str(ecif_copy))
        assert result["sheet_count"] > 5
        assert result["has_vba"] is True

        # Check for known sheets
        names = [s["name"] for s in result["sheets"]]
        assert "ECIF Work Scope (E)" in names

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_list_ecif_tables(self, excel_advanced_tools, ecif_copy):
        """Should find Milestones table."""
        result = excel_advanced_tools.tool_excel_list_tables(str(ecif_copy))
        names = [t["name"] for t in result["tables"]]
        assert "Milestones" in names

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_get_milestones_table(self, excel_advanced_tools, ecif_copy):
        """Should read Milestones table data."""
        result = excel_advanced_tools.tool_excel_get_table(str(ecif_copy), "Milestones")
        assert result["table_name"] == "Milestones"
        assert "Milestone #" in result["columns"]

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_list_ecif_named_ranges(self, excel_advanced_tools, ecif_copy):
        """Should list ECIF named ranges."""
        result = excel_advanced_tools.tool_excel_list_named_ranges(str(ecif_copy))
        # ECIF has many named ranges for dropdowns
        assert result["count"] > 10

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_patch_ecif_cell(self, excel_advanced_tools, ecif_copy):
        """Should patch cell and preserve VBA."""
        result = excel_advanced_tools.tool_excel_patch_cell(
            str(ecif_copy),
            "B5",
            "Contoso Corporation",
            sheet_name="ECIF Work Scope (E)",
        )
        assert result["success"] is True

        # Verify file is still valid XLSM
        from openpyxl import load_workbook

        wb = load_workbook(ecif_copy, keep_vba=True)
        assert wb["ECIF Work Scope (E)"]["B5"].value == "Contoso Corporation"
        wb.close()


# =============================================================================
# EDGE CASES AND ERROR HANDLING
# =============================================================================


class TestEdgeCases:
    """Tests for edge cases and error conditions."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_empty_workbook(self, excel_advanced_tools, temp_dir):
        """Should handle empty workbook."""
        wb = Workbook()
        path = temp_dir / "empty.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_list_sheets(str(path))
        assert result["sheet_count"] == 1

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_unicode_content(self, excel_advanced_tools, temp_dir):
        """Should handle unicode characters."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "日本語テスト"
        ws["A2"] = "Ñoño"
        path = temp_dir / "unicode.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_get_range(str(path), "A1:A2")
        assert "日本語テスト" in str(result["data"])

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_special_characters_in_sheet_name(self, excel_advanced_tools, temp_dir):
        """Should handle special characters in sheet names."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data (2026)"
        ws["A1"] = "Test"
        path = temp_dir / "special.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_get_range(
            str(path), "A1", sheet_name="Data (2026)"
        )
        assert result["data"][0][0]["value"] == "Test"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_large_range(self, excel_advanced_tools, temp_dir):
        """Should handle moderately large ranges."""
        wb = Workbook()
        ws = wb.active
        for row in range(1, 101):
            for col in range(1, 11):
                ws.cell(row=row, column=col, value=f"R{row}C{col}")
        path = temp_dir / "large.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_get_range(str(path), "A1:J100")
        assert result["row_count"] == 100
        assert result["col_count"] == 10

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_numeric_coercion(self, excel_advanced_tools, temp_dir):
        """Should preserve numeric types."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 123
        ws["A2"] = 45.67
        ws["A3"] = "Not a number"
        path = temp_dir / "numbers.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_get_range(str(path), "A1:A3")
        assert result["data"][0][0]["type"] == "int"
        assert result["data"][1][0]["type"] == "float"
        assert result["data"][2][0]["type"] == "str"


# =============================================================================
# NEW TESTS: Merged Cells
# =============================================================================


class TestListMergedCells:
    """Tests for tool_excel_list_merged_cells."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_lists_merged_cells(self, excel_advanced_tools, temp_dir):
        """Should list all merged cell regions."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.merge_cells("A1:C1")
        ws.merge_cells("B3:B5")
        ws["A1"] = "Header spanning 3 columns"
        path = temp_dir / "merged.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_list_merged_cells(str(path))
        assert result["total_merged_regions"] == 2
        assert "Data" in result["by_sheet"]
        ranges = [m["range"] for m in result["by_sheet"]["Data"]]
        assert "A1:C1" in ranges
        assert "B3:B5" in ranges

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_filter_by_sheet(self, excel_advanced_tools, temp_dir):
        """Should filter merged cells by sheet name."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.merge_cells("A1:B1")

        ws2 = wb.create_sheet("Sheet2")
        ws2.merge_cells("C1:D1")

        path = temp_dir / "multi_merge.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_list_merged_cells(str(path), sheet_name="Sheet1")
        assert result["total_merged_regions"] == 1
        assert "Sheet1" in result["by_sheet"]
        assert "Sheet2" not in result["by_sheet"]

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_no_merged_cells(self, excel_advanced_tools, temp_dir):
        """Should handle workbooks with no merged cells."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "No merges"
        path = temp_dir / "no_merge.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_list_merged_cells(str(path))
        assert result["total_merged_regions"] == 0
        assert result["by_sheet"] == {}

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_list_merged_cells("/nonexistent.xlsx")
        assert "error" in result

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_includes_top_left_cell(self, excel_advanced_tools, temp_dir):
        """Should include top_left cell for easy reference."""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("B2:D4")
        path = temp_dir / "topleft.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_list_merged_cells(str(path))
        merged = result["by_sheet"]["Sheet"][0]
        assert merged["top_left"] == "B2"
        assert merged["min_row"] == 2
        assert merged["max_row"] == 4


# =============================================================================
# NEW TESTS: Comments
# =============================================================================


class TestAddComment:
    """Tests for tool_excel_add_comment."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_adds_comment(self, excel_advanced_tools, temp_dir):
        """Should add a comment to a cell."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Data"
        path = temp_dir / "comment.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_add_comment(
            str(path), "A1", "This is a test comment", author="Test User"
        )
        assert result["success"] is True
        assert result["appended"] is False
        assert result["author"] == "Test User"

        # Verify comment exists
        from openpyxl import load_workbook

        wb = load_workbook(path)
        assert wb.active["A1"].comment is not None
        assert "test comment" in wb.active["A1"].comment.text
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_appends_to_existing_comment(self, excel_advanced_tools, temp_dir):
        """Should append to existing comment."""
        from openpyxl.comments import Comment

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Data"
        ws["A1"].comment = Comment("Original comment", "Author1")
        path = temp_dir / "append_comment.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_add_comment(
            str(path), "A1", "Appended text", author="Author2"
        )
        assert result["success"] is True
        assert result["appended"] is True

        # Verify both comments present
        from openpyxl import load_workbook

        wb = load_workbook(path)
        comment_text = wb.active["A1"].comment.text
        assert "Original comment" in comment_text
        assert "Appended text" in comment_text
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_sheet_in_reference(self, excel_advanced_tools, temp_dir):
        """Should handle sheet name in cell reference."""
        wb = Workbook()
        ws = wb.active
        ws.title = "MySheet"
        ws["B5"] = "Value"
        path = temp_dir / "sheet_ref.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_add_comment(
            str(path), "MySheet!B5", "Comment via sheet ref"
        )
        assert result["success"] is True
        assert result["sheet"] == "MySheet"

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_add_comment(
            "/nonexistent.xlsx", "A1", "Comment"
        )
        assert "error" in result

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_sheet_not_found(self, excel_advanced_tools, temp_dir):
        """Should handle missing sheet."""
        wb = Workbook()
        path = temp_dir / "no_sheet.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_add_comment(
            str(path), "A1", "Comment", sheet_name="NonExistent"
        )
        assert "error" in result


class TestGetComments:
    """Tests for tool_excel_get_comments."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_gets_comments(self, excel_advanced_tools, temp_dir):
        """Should retrieve all comments from workbook."""
        from openpyxl.comments import Comment

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Value 1"
        ws["A1"].comment = Comment("Comment 1", "Author1")
        ws["B2"] = "Value 2"
        ws["B2"].comment = Comment("Comment 2", "Author2")
        path = temp_dir / "get_comments.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_get_comments(str(path))
        assert result["total_comments"] == 2
        assert "Data" in result["by_sheet"]

        comments = result["by_sheet"]["Data"]
        cells = [c["cell"] for c in comments]
        assert "A1" in cells
        assert "B2" in cells

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_filter_by_sheet(self, excel_advanced_tools, temp_dir):
        """Should filter comments by sheet name."""
        from openpyxl.comments import Comment

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"].comment = Comment("Comment in Sheet1", "Author")

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"].comment = Comment("Comment in Sheet2", "Author")

        path = temp_dir / "multi_comments.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_get_comments(str(path), sheet_name="Sheet1")
        assert result["total_comments"] == 1
        assert "Sheet1" in result["by_sheet"]
        assert "Sheet2" not in result["by_sheet"]

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_no_comments(self, excel_advanced_tools, temp_dir):
        """Should handle workbooks with no comments."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "No comments"
        path = temp_dir / "no_comments.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_get_comments(str(path))
        assert result["total_comments"] == 0
        assert result["by_sheet"] == {}

    def test_file_not_found(self, excel_advanced_tools):
        """Should handle missing files."""
        result = excel_advanced_tools.tool_excel_get_comments("/nonexistent.xlsx")
        assert "error" in result

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_includes_cell_value(self, excel_advanced_tools, temp_dir):
        """Should include the cell value with comment."""
        from openpyxl.comments import Comment

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Important Data"
        ws["A1"].comment = Comment("Note about this", "Author")
        path = temp_dir / "value_comment.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_get_comments(str(path))
        comment = result["by_sheet"]["Sheet"][0]
        assert comment["cell_value"] == "Important Data"
        assert comment["author"] == "Author"


class TestDeleteComments:
    """Tests for tool_excel_delete_comment."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_deletes_comment(self, excel_advanced_tools, temp_dir):
        """Should delete an existing cell comment."""
        from openpyxl.comments import Comment

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Data"
        ws["A1"].comment = Comment("To delete", "Author")
        path = temp_dir / "delete_comment.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_delete_comment(str(path), "A1")
        assert result.get("success") is True

        from openpyxl import load_workbook
        wb2 = load_workbook(path)
        assert wb2.active["A1"].comment is None
        wb2.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_delete_missing_comment_errors(self, excel_advanced_tools, temp_dir):
        """Should return an error when no comment exists on target cell."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "No comment"
        path = temp_dir / "delete_missing_comment.xlsx"
        wb.save(path)

        result = excel_advanced_tools.tool_excel_delete_comment(str(path), "A1")
        assert "error" in result


# =============================================================================
# NEW TESTS: Highlight Default Changed
# =============================================================================


class TestHighlightDefault:
    """Tests to verify highlight defaults to False."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_patch_cell_no_highlight_by_default(self, excel_advanced_tools, temp_dir):
        """patch_cell should not highlight by default (preserves formatting)."""
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Original"
        # Set a custom fill color
        ws["A1"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        path = temp_dir / "no_highlight.xlsx"
        wb.save(path)

        # Patch without explicit highlight parameter (should default to False)
        excel_advanced_tools.tool_excel_patch_cell(str(path), "A1", "Updated")

        # Verify fill is preserved (not changed to yellow)
        from openpyxl import load_workbook

        wb = load_workbook(path)
        cell_fill = wb.active["A1"].fill.start_color.rgb
        # Should still be red, not the highlight yellow (FFFF99)
        assert cell_fill != "00FFFF99"
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_patch_cell_explicit_highlight(self, excel_advanced_tools, temp_dir):
        """patch_cell should highlight when explicitly requested."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Original"
        path = temp_dir / "explicit_highlight.xlsx"
        wb.save(path)

        # Patch with explicit highlight=True
        excel_advanced_tools.tool_excel_patch_cell(str(path), "A1", "Updated", highlight=True)

        # Verify fill is yellow highlight
        from openpyxl import load_workbook

        wb = load_workbook(path)
        cell_fill = wb.active["A1"].fill.start_color.rgb
        assert cell_fill == "00FFFF99"
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_patch_range_no_highlight_by_default(self, excel_advanced_tools, temp_dir):
        """patch_range should not highlight by default."""
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "A"
        ws["B1"] = "B"
        ws["A1"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        path = temp_dir / "range_no_highlight.xlsx"
        wb.save(path)

        # Patch range without explicit highlight
        excel_advanced_tools.tool_excel_patch_range(str(path), "A1:B1", [["X", "Y"]])

        # Verify fill is preserved
        from openpyxl import load_workbook

        wb = load_workbook(path)
        cell_fill = wb.active["A1"].fill.start_color.rgb
        assert cell_fill != "00FFFF99"
        wb.close()
