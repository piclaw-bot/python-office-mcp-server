"""
Tests for excel_tools.py - Excel spreadsheet processing

Tests cover:
- Excel extraction
- Markdown to Excel conversion
- Excel to Markdown conversion
"""

import pytest

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False



# Fixtures temp_dir, excel_tools, and sample_xlsx are provided by conftest.py


class TestExcelExtract:
    """Tests for tool_excel_extract."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_extracts_data(self, excel_tools, sample_xlsx):
        """Should extract data from Excel file."""
        result = excel_tools.tool_excel_extract(str(sample_xlsx))
        assert "sheets" in result
        assert len(result["sheets"]) >= 1

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_extracts_headers(self, excel_tools, sample_xlsx):
        """Should extract column headers."""
        result = excel_tools.tool_excel_extract(str(sample_xlsx))
        sheets = result.get("sheets", {})
        # sheets is a dict of sheet_name -> rows
        if sheets:
            first_sheet = list(sheets.values())[0]
            first_row = first_sheet[0] if first_sheet else []
            assert "Name" in first_row

    def test_file_not_found(self, excel_tools):
        """Should handle missing files."""
        result = excel_tools.tool_excel_extract("/nonexistent.xlsx")
        assert "error" in result


class TestExcelToMarkdown:
    """Tests for tool_excel_to_markdown."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_converts_to_markdown(self, excel_tools, sample_xlsx):
        """Should convert Excel to markdown tables."""
        result = excel_tools.tool_excel_to_markdown(str(sample_xlsx))
        # Returns string directly
        assert isinstance(result, str)
        assert "|" in result
        assert "Name" in result

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_includes_all_columns(self, excel_tools, sample_xlsx):
        """Should include all columns in markdown."""
        result = excel_tools.tool_excel_to_markdown(str(sample_xlsx))
        assert "Name" in result
        assert "Value" in result
        assert "Status" in result

    def test_file_not_found(self, excel_tools):
        """Should handle missing files."""
        result = excel_tools.tool_excel_to_markdown("/nonexistent.xlsx")
        # Returns error string
        assert "Error" in result or "not found" in result.lower()


class TestExcelFromMarkdown:
    """Tests for tool_excel_from_markdown."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_creates_excel(self, excel_tools, temp_dir):
        """Should create Excel file from markdown table."""
        md = """
| Name | Value | Status |
|------|-------|--------|
| A    | 100   | Active |
| B    | 200   | Done   |
"""
        output = temp_dir / "created.xlsx"
        result = excel_tools.tool_excel_from_markdown(str(output), md)
        assert result.get("success") is True
        assert output.exists()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_creates_excel_from_markdown_file(self, excel_tools, temp_dir):
        """Should create Excel from markdown_file path."""
        md_file = temp_dir / "input.md"
        md_file.write_text(
            """
| Name | Value |
|------|-------|
| A    | 100   |
""",
            encoding="utf-8",
        )

        output = temp_dir / "from_file.xlsx"
        result = excel_tools.tool_excel_from_markdown(
            str(output),
            markdown_file=str(md_file),
        )
        assert result.get("success") is True
        assert output.exists()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_preserves_headers(self, excel_tools, temp_dir):
        """Should preserve column headers."""
        md = """
| Col1 | Col2 |
|------|------|
| A    | B    |
"""
        output = temp_dir / "headers.xlsx"
        excel_tools.tool_excel_from_markdown(str(output), md)

        wb = Workbook()
        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb.active
        assert ws['A1'].value == "Col1"
        assert ws['B1'].value == "Col2"
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_multiple_tables(self, excel_tools, temp_dir):
        """Should handle multiple tables as separate sheets."""
        md = """
| Table1 | Data |
|--------|------|
| A      | 1    |

| Table2 | Info |
|--------|------|
| X      | Y    |
"""
        output = temp_dir / "multi.xlsx"
        result = excel_tools.tool_excel_from_markdown(str(output), md)
        assert result.get("success") is True

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_sheet_names_from_headings(self, excel_tools, temp_dir):
        """Should name sheets from preceding H2 headings."""
        md = """
## First Section

| Col1 | Col2 |
|------|------|
| A    | B    |

## Second Section

| Col1 | Col2 |
|------|------|
| C    | D    |
"""
        output = temp_dir / "headings.xlsx"
        excel_tools.tool_excel_from_markdown(str(output), md)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert wb.sheetnames[0] == "First Section"
        assert wb.sheetnames[1] == "Second Section"
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_formula_cells(self, excel_tools, temp_dir):
        """Should preserve formulas when cell starts with '='."""
        md = """
| Value | Calc |
|-------|------|
| 10    | =A2*2 |
"""
        output = temp_dir / "formula.xlsx"
        excel_tools.tool_excel_from_markdown(str(output), md)

        from openpyxl import load_workbook
        wb = load_workbook(output, data_only=False)
        ws = wb.active
        assert ws["B2"].value == "=A2*2"
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_sheet_name_parameter(self, excel_tools, temp_dir):
        """Should apply sheet_name to the first sheet."""
        md = """
| Col1 | Col2 |
|------|------|
| A    | B    |

| Col1 | Col2 |
|------|------|
| C    | D    |
"""
        output = temp_dir / "named.xlsx"
        excel_tools.tool_excel_from_markdown(str(output), md, sheet_name="Primary")

        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert wb.sheetnames[0] == "Primary"
        wb.close()


class TestListSupportedFormats:
    """Tests for list_supported_formats."""

    def test_tool_exists(self, excel_tools):
        """ExcelTools should have expected methods."""
        # Check the tool has expected methods
        assert hasattr(excel_tools, 'tool_excel_extract')
        assert hasattr(excel_tools, 'tool_excel_to_markdown')
        assert hasattr(excel_tools, 'tool_excel_from_markdown')
