"""Tests for Excel chart tooling."""

import pytest

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class TestExcelCharts:
    """Tests for tool_excel_add_chart."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_adds_chart(self, excel_advanced_tools, temp_dir):
        """Should add a chart to the worksheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Month", "Value"])
        ws.append(["Jan", 10])
        ws.append(["Feb", 20])
        ws.append(["Mar", 30])

        path = temp_dir / "chart.xlsx"
        wb.save(path)
        wb.close()

        result = excel_advanced_tools.tool_excel_add_chart(
            file_path=str(path),
            data_range="Data!A1:B4",
            chart_type="line",
            title="Trend",
            position="E2",
        )

        assert result.get("success") is True

        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["Data"]
        assert len(ws._charts) == 1
        wb.close()
