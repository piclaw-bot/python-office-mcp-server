#!/usr/bin/env python3
"""
excel_tools.py - MCP tools for Excel workbook processing

Provides tools to parse, extract content from, and generate Excel (.xlsx) files.
Supports GitHub Flavored Markdown table parsing to extract tables into separate sheets.
"""

import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .markdown_parser import (
    Paragraph,
    Table,
    extract_tables_from_markdown,
    parse_markdown_to_nodes,
)
from .save_utils import resolve_office_path


class ExcelTools:
    """MCP tool mixin for Excel workbook processing.

    Note: The extract/to_markdown methods are kept for internal use by
    office_unified_tools but are not exposed as separate MCP tools.
    They are filtered out in office_server.py via DEPRECATED_TOOLS.

    Public MCP tools:
    - excel_from_markdown: Convert Markdown tables to Excel workbook

    Internal methods (used by unified tools):
    - tool_excel_extract: Extract data from workbook
    - tool_excel_to_markdown: Convert sheets to Markdown
    """

    def tool_excel_extract(self, file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
        """Extract data from an Excel workbook.

        NOTE: This method is internal - use office_read() instead.

        Args:
            file_path: Path to the .xlsx file
            sheet_name: Optional specific sheet name (extracts all if not provided)

        Returns:
            Dictionary with sheet names and their data as lists of rows
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        wb = load_workbook(file_path, data_only=True)
        try:
            result = {"file": path.name, "sheets": {}}

            sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames

            for name in sheets_to_process:
                if name not in wb.sheetnames:
                    continue
                ws = wb[name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    # Convert to strings, handle None
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_data):  # Skip completely empty rows
                        rows.append(row_data)
                result["sheets"][name] = rows

            return result
        finally:
            wb.close()

    def tool_excel_to_markdown(self, file_path: str, sheet_name: str | None = None) -> str:
        """Convert Excel sheets to Markdown tables.

        NOTE: This method is internal - use office_read(output_format="markdown") instead.

        Args:
            file_path: Path to the .xlsx file
            sheet_name: Optional specific sheet name

        Returns:
            Markdown string with tables for each sheet
        """
        data = self.tool_excel_extract(file_path, sheet_name)
        if "error" in data:
            return f"Error: {data['error']}"

        lines = []
        for name, rows in data["sheets"].items():
            lines.append(f"## {name}")
            lines.append("")

            if not rows:
                lines.append("*Empty sheet*")
                lines.append("")
                continue

            # Header
            header = rows[0]
            lines.append("| " + " | ".join(header) + " |")
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")

            # Data
            for row in rows[1:]:
                # Pad row if needed
                while len(row) < len(header):
                    row.append("")
                lines.append("| " + " | ".join(row[:len(header)]) + " |")
            lines.append("")

        return "\n".join(lines)

    def tool_excel_from_markdown(
        self,
        output_path: str,
        markdown: str | None = None,
        sheet_name: str | None = None,
        markdown_file: str | None = None,
    ) -> dict[str, Any]:
        """Convert Markdown tables to an Excel workbook.

        This is the primary tool for creating Excel workbooks from text content.

        Parses GitHub Flavored Markdown content and extracts all tables.
        Each table becomes a separate sheet in the workbook.

        Features:
        - Auto-detects multiple tables in the content using GFM parser
        - Header row gets bold formatting with gray background
        - Auto-filter on header row
        - Column widths auto-sized based on content
        - Numbers and percentages are coerced to numeric types
        - Supports inline formatting in cells (bold, italic, code)
        - Supports formulas when a cell starts with '='
        - Uses nearby '##' headings to name sheets (up to 31 chars)

        Example:
            excel_from_markdown(
                output_path="04. Artifacts/budget.xlsx",
                markdown='''
| Category | Q1 | Q2 | Q3 | Q4 | Total |
|----------|----|----|----|----|-------|
| Personnel | $50,000 | $52,000 | $54,000 | $56,000 | $212,000 |
| Infrastructure | $15,000 | $15,000 | $16,000 | $16,000 | $62,000 |
| Software | $8,000 | $8,500 | $9,000 | $9,500 | $35,000 |
| Training | $5,000 | $3,000 | $4,000 | $3,000 | $15,000 |

| Milestone | Target Date | Owner | Status |
|-----------|-------------|-------|--------|
| Phase 1 Complete | 2026-03-31 | Alice | 100% |
| Phase 2 Complete | 2026-06-30 | Bob | 45% |
| Go-Live | 2026-09-30 | Carol | 0% |
'''
            )

        Args:
            output_path: Path for the output .xlsx file
            markdown: GitHub Flavored Markdown content containing one or more tables (inline)
            sheet_name: Optional sheet name for the first/only sheet
            markdown_file: Optional path to a Markdown file. Use this for
                very large inputs to avoid MCP argument-size limits.

        Returns:
            Status dictionary with file path and sheet count
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        markdown_text = markdown
        if markdown_file:
            resolved_md_path = resolve_office_path(markdown_file)
            md_path = Path(resolved_md_path)
            if not md_path.exists():
                return {"error": f"Markdown file not found: {markdown_file}"}
            if not md_path.is_file():
                return {"error": f"Markdown path is not a file: {markdown_file}"}
            markdown_text = md_path.read_text(encoding="utf-8")

        if markdown_text is None:
            return {"error": "Provide either 'markdown' or 'markdown_file'"}

        # Parse all tables from markdown using GFM parser
        tables = extract_tables_from_markdown(markdown_text)

        if not tables:
            return {"error": "No Markdown tables found. Use pipe (|) format: | col1 | col2 |"}

        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        nodes = parse_markdown_to_nodes(markdown_text)
        tables_with_headings: list[tuple[Table, str | None]] = []
        current_heading: str | None = None
        for node in nodes:
            if isinstance(node, Paragraph) and node.level == 2:
                heading_text = node.text.strip()
                if heading_text:
                    current_heading = heading_text
            elif isinstance(node, Table):
                tables_with_headings.append((node, current_heading))

        if not tables_with_headings:
            return {"error": "No Markdown tables found. Use pipe (|) format: | col1 | col2 |"}

        existing_names: set[str] = set()
        sheet_index = 1
        for table, heading in tables_with_headings:
            # Determine sheet name
            if sheet_index == 1 and sheet_name:
                base_name = sheet_name
            elif heading:
                base_name = heading
            else:
                base_name = f"Sheet{sheet_index}"

            sheet_index += 1
            clean_name = self._sanitize_sheet_name(base_name)
            if not clean_name:
                clean_name = f"Sheet{sheet_index - 1}"
            final_name = self._dedupe_sheet_name(clean_name, existing_names)
            existing_names.add(final_name)

            ws = wb.create_sheet(title=final_name)

            # Convert Table object to row data
            all_rows = []
            for row in table.rows:
                row_data = [cell.text for cell in row.cells]
                all_rows.append(row_data)

            for row_idx, row_data in enumerate(all_rows, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)

                    value_text = str(value)
                    if value_text.strip().startswith("="):
                        cell.value = value_text.strip()
                        continue

                    # Try to coerce numbers (including currency)
                    coerced = self._coerce_number(value_text)
                    if coerced is not None:
                        cell.value = coerced["value"]
                        if coerced.get("format"):
                            cell.number_format = coerced["format"]
                    else:
                        cell.value = value_text

            # Format header row (first row) - bold with light gray background
            if all_rows:
                for col_idx in range(1, len(all_rows[0]) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

                # Auto-fit columns (approximate, since openpyxl doesn't have auto-fit)
                for col_idx in range(1, len(all_rows[0]) + 1):
                    max_width = 8
                    for row in all_rows:
                        if col_idx <= len(row):
                            cell_len = len(str(row[col_idx - 1]))
                            max_width = max(max_width, min(50, cell_len + 2))
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_width

                # Add auto-filter on data range
                ws.auto_filter.ref = f"A1:{get_column_letter(len(all_rows[0]))}{len(all_rows)}"

        wb.save(output_path)
        return {
            "success": True,
            "file": output_path,
            "sheets": len(tables_with_headings),
            "message": f"Created Excel workbook with {len(tables_with_headings)} sheet(s)"
        }

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize a worksheet name to meet Excel constraints."""
        cleaned = re.sub(r"[\\/*?:\[\]]", " ", name).strip()
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned[:31]

    def _dedupe_sheet_name(self, name: str, existing: set[str]) -> str:
        """Ensure the worksheet name is unique within the workbook."""
        if name not in existing:
            return name

        suffix = 2
        while True:
            candidate = f"{name} ({suffix})"
            candidate = candidate[:31]
            if candidate not in existing:
                return candidate
            suffix += 1

    def _coerce_number(self, value: str) -> dict[str, Any] | None:
        """Try to coerce a string value to a number.

        Handles:
        - Plain numbers (with optional thousands separators)
        - Percentages (e.g., "45%" -> 0.45 with 0% format)
        - Currency (e.g., "$1,234.56" -> 1234.56 with $#,##0.00 format)

        Returns dict with 'value' and optional 'format', or None if not numeric.
        """
        s = str(value).strip()
        if not s:
            return None

        # Check for currency ($ prefix)
        is_currency = s.startswith('$')
        if is_currency:
            s = s[1:]  # Remove $ for parsing

        # Remove thousands separators
        normalized = s.replace(",", "")

        # Percentage
        if normalized.endswith('%'):
            try:
                n = float(normalized[:-1]) / 100
                return {"value": n, "format": "0%"}
            except ValueError:
                return None

        # Plain number (or currency number)
        if re.match(r'^-?\d+(?:\.\d+)?$', normalized):
            try:
                n = float(normalized)
                if is_currency:
                    return {"value": n, "format": "$#,##0.00"}
                # Use int if it's a whole number
                if n == int(n):
                    return {"value": int(n)}
                return {"value": n}
            except ValueError:
                return None

        return None
