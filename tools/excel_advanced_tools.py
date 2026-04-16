#!/usr/bin/env python3
"""
excel_advanced_tools.py - Advanced MCP tools for Excel workbook manipulation

Provides comprehensive tools to:
- Introspect workbook structure (sheets, tables, named ranges)
- Read and patch cells/ranges with change logging
- Work with Excel tables (ListObjects)
- Fill templates with placeholder replacement
- Audit workbooks for unfilled placeholders

Quality bar: All tools preserve workbook structure including VBA (for XLSM),
formulas, formatting, and validation rules.
"""

import contextlib
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

try:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.utils.cell import range_boundaries

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


from .diagnostics import build_mutation_diagnostics

# Author name for change tracking (from environment or default)
DEFAULT_AUTHOR = os.environ.get("MCP_AUTHOR", "Solution Architect Agent")

# Change highlight color (light yellow)
CHANGE_HIGHLIGHT_COLOR = "FFFF99"


def _ensure_change_log_sheet(wb) -> None:
    """Ensure the workbook has a _ChangeLog sheet for tracking changes."""
    if "_ChangeLog" not in wb.sheetnames:
        log_ws = wb.create_sheet("_ChangeLog")
        log_ws.append(["Timestamp", "Sheet", "Cell", "Old Value", "New Value", "Author"])
        # Format header row
        for cell in log_ws[1]:
            cell.font = Font(bold=True)
        # Set column widths
        log_ws.column_dimensions["A"].width = 22
        log_ws.column_dimensions["B"].width = 20
        log_ws.column_dimensions["C"].width = 10
        log_ws.column_dimensions["D"].width = 30
        log_ws.column_dimensions["E"].width = 30
        log_ws.column_dimensions["F"].width = 25


def _log_change(wb, sheet_name: str, cell_ref: str, old_value, new_value, author: str) -> None:
    """Log a change to the _ChangeLog sheet."""
    _ensure_change_log_sheet(wb)
    log_ws = wb["_ChangeLog"]
    log_ws.append([
        datetime.now().isoformat(),
        sheet_name,
        cell_ref,
        str(old_value) if old_value is not None else "",
        str(new_value) if new_value is not None else "",
        author,
    ])


def _highlight_cell(cell) -> None:
    """Apply visual highlight to a changed cell."""
    cell.fill = PatternFill(
        start_color=CHANGE_HIGHLIGHT_COLOR,
        end_color=CHANGE_HIGHLIGHT_COLOR,
        fill_type="solid",
    )


def _add_change_comment(cell, old_value, new_value, author: str) -> None:
    """Add or update a comment documenting the change."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    text = f"Changed by {author} at {timestamp}\nOld: {old_value}\nNew: {new_value}"

    if cell.comment:
        cell.comment.text += f"\n\n{text}"
    else:
        cell.comment = Comment(text, author)
        cell.comment.width = 250
        cell.comment.height = 100


def _parse_cell_reference(ref: str) -> tuple[str | None, str]:
    """Parse a cell reference, optionally with sheet name.

    Args:
        ref: Cell reference like 'A1', 'B5', or 'Sheet1!A1'

    Returns:
        Tuple of (sheet_name or None, cell_ref)
    """
    if "!" in ref:
        parts = ref.split("!", 1)
        sheet = parts[0].strip("'\"")
        cell = parts[1]
        return sheet, cell
    return None, ref


def _parse_range_reference(ref: str) -> tuple[str | None, str]:
    """Parse a range reference, optionally with sheet name.

    Args:
        ref: Range reference like 'A1:D10' or 'Sheet1!A1:D10'

    Returns:
        Tuple of (sheet_name or None, range_ref)
    """
    if "!" in ref:
        parts = ref.split("!", 1)
        sheet = parts[0].strip("'\"")
        range_ref = parts[1]
        return sheet, range_ref
    return None, ref


def _get_range_bounds(range_ref: str) -> tuple[int, int, int, int]:
    """Parse a range reference like 'A1:D10' into bounds.

    Returns:
        Tuple of (min_row, min_col, max_row, max_col)
    """
    if ":" in range_ref:
        start, end = range_ref.split(":")
    else:
        start = end = range_ref

    # Parse start
    start_match = re.match(r"([A-Z]+)(\d+)", start.upper())
    end_match = re.match(r"([A-Z]+)(\d+)", end.upper())

    if not start_match or not end_match:
        raise ValueError(f"Invalid range reference: {range_ref}")

    min_col = column_index_from_string(start_match.group(1))
    min_row = int(start_match.group(2))
    max_col = column_index_from_string(end_match.group(1))
    max_row = int(end_match.group(2))

    return min_row, min_col, max_row, max_col


def _coerce_value(value: Any) -> Any:
    """Coerce string values to appropriate types (numbers, dates) when possible.

    Preserves formulas (starting with '=') and handles common numeric formats
    including currency symbols, percentages, and thousands separators.

    Args:
        value: The value to potentially convert

    Returns:
        The coerced value (int, float, or original value if no conversion applies)
    """
    if value is None:
        return None

    # Already the right type
    if isinstance(value, (int, float, bool)) and not isinstance(value, str):
        return value

    if not isinstance(value, str):
        return value

    text = value.strip()

    # Empty string
    if not text:
        return value

    # Preserve formulas
    if text.startswith("="):
        return value

    # Try to parse as number
    # Remove common formatting: currency symbols, thousands separators
    cleaned = text

    # Remove currency symbols (common ones)
    for symbol in ["$", "€", "£", "¥", "kr", "NOK", "USD", "EUR"]:
        cleaned = cleaned.replace(symbol, "").strip()

    # Handle percentage
    is_percentage = cleaned.endswith("%")
    if is_percentage:
        cleaned = cleaned[:-1].strip()

    # Remove thousands separators (be careful with decimal points)
    # Common patterns: 1,000.00 (US) or 1.000,00 (EU)
    # Heuristic: if we have both . and , check which is the decimal
    if "," in cleaned and "." in cleaned:
        # If comma comes after period, comma is decimal (EU style)
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            # Period is decimal (US style)
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        # Could be thousands separator or decimal
        # If exactly 3 digits after last comma, treat as thousands sep
        parts = cleaned.split(",")
        if len(parts[-1]) == 3 and parts[-1].isdigit():
            cleaned = cleaned.replace(",", "")
        else:
            # Treat as decimal separator
            cleaned = cleaned.replace(",", ".")

    # Remove leading/trailing whitespace from cleaned
    cleaned = cleaned.strip()

    # Try integer first
    try:
        int_val = int(cleaned)
        if is_percentage:
            return int_val / 100.0
        return int_val
    except ValueError:
        pass

    # Try float
    try:
        float_val = float(cleaned)
        if is_percentage:
            return float_val / 100.0
        return float_val
    except ValueError:
        pass

    # Return original value if no conversion
    return value


def _auto_row_height(ws, row: int, cell=None, default_height: float = 15.0) -> None:
    """Auto-adjust row height for cells with multi-line content or text wrap.

    Calculates appropriate height based on:
    - Number of newlines in cell content
    - Approximate text wrap based on column width
    - Minimum height constraints

    Args:
        ws: The worksheet object
        row: The row number (1-based)
        cell: Optional specific cell to check (otherwise checks all cells in row)
        default_height: Default row height in points (15.0 is Excel's default)
    """
    # Constants for height calculation
    LINE_HEIGHT = 15.0  # Points per line of text
    MIN_HEIGHT = 15.0   # Minimum row height
    MAX_HEIGHT = 409.0  # Excel's maximum row height

    max_lines = 1

    if cell is not None:
        cells_to_check = [cell]
    else:
        # Check all cells in the row that have values
        cells_to_check = [c for c in ws[row] if c.value is not None]

    for c in cells_to_check:
        if c.value is None:
            continue

        text = str(c.value)

        # Estimate wrapped lines based on column width
        col_letter = get_column_letter(c.column)
        col_width = ws.column_dimensions[col_letter].width or 8.43  # Default Excel width

        # Approximate characters per line (col_width is in character units)
        chars_per_line = max(1, int(col_width))

        # Calculate wrapped lines for each explicit line
        total_lines = 0
        for line in text.split("\n"):
            line_chars = len(line)
            wrapped = max(1, (line_chars + chars_per_line - 1) // chars_per_line)
            total_lines += wrapped

        max_lines = max(max_lines, total_lines)

    # Calculate required height
    required_height = max_lines * LINE_HEIGHT

    # Apply constraints
    required_height = max(MIN_HEIGHT, min(MAX_HEIGHT, required_height))

    # Only adjust if new height is greater than current
    current_height = ws.row_dimensions[row].height or default_height
    if required_height > current_height:
        ws.row_dimensions[row].height = required_height


def _set_cell_with_coercion(cell, value: Any, auto_height: bool = True) -> Any:
    """Set a cell value with type coercion and optional row height adjustment.

    Args:
        cell: The openpyxl cell object
        value: The value to set
        auto_height: Whether to enable text wrap for multi-line content

    Returns:
        The coerced value that was set
    """
    from openpyxl.styles import Alignment

    coerced = _coerce_value(value)
    cell.value = coerced

    # Enable text wrap for multi-line content
    if auto_height and isinstance(coerced, str) and "\n" in coerced:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    return coerced


def _close_workbook(wb) -> None:
    """Close an openpyxl workbook and any VBA archive safely."""
    vba_archive = getattr(wb, "vba_archive", None)
    if vba_archive is not None:
        with contextlib.suppress(Exception):
            vba_archive.close()
        with contextlib.suppress(Exception):
            wb.vba_archive = None

    archive = getattr(wb, "_archive", None)
    if archive is not None:
        with contextlib.suppress(Exception):
            archive.close()
        with contextlib.suppress(Exception):
            wb._archive = None

    with contextlib.suppress(Exception):
        wb.close()


def _load_workbook_for_path(file_path: str, **kwargs):
    """Load a workbook, enabling VBA preservation only for macro-enabled formats.

    Passing keep_vba=True for plain .xlsx files can leave zip-backed workbook
    state in a noisier teardown path for otherwise simple/empty workbooks.
    Restrict it to macro-enabled extensions where preservation is actually needed.
    """
    suffix = Path(file_path).suffix.lower()
    kwargs.setdefault("keep_vba", suffix in {".xlsm", ".xltm"})
    return load_workbook(file_path, **kwargs)


class ExcelAdvancedTools:
    """MCP tool mixin for advanced Excel workbook manipulation."""

    # =========================================================================
    # WORKBOOK INTROSPECTION
    # =========================================================================

    def tool_excel_list_sheets(
        self,
        file_path: str,
        include_hidden: bool = True,
    ) -> dict[str, Any]:
        """List all sheets in an Excel workbook with their properties.

        Provides detailed information about each sheet including visibility,
        dimensions, and whether it contains tables or data validations.

        Example:
            excel_list_sheets(file_path="template.xlsx")

        Args:
            file_path: Path to the .xlsx or .xlsm file
            include_hidden: Include hidden and very hidden sheets (default True)

        Returns:
            Dictionary with sheet information including names, states, and metadata
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                state = ws.sheet_state  # visible, hidden, veryHidden

                if not include_hidden and state != "visible":
                    continue

                sheet_info = {
                    "name": name,
                    "state": state,
                    "dimensions": ws.dimensions or "Empty",
                    "tables": list(ws.tables.keys()) if ws.tables else [],
                    "has_data_validations": bool(
                        ws.data_validations and ws.data_validations.dataValidation
                    ),
                    "merged_cell_count": len(ws.merged_cells.ranges),
                }
                sheets.append(sheet_info)

            return {
                "file": path.name,
                "sheet_count": len(sheets),
                "sheets": sheets,
                "has_vba": path.suffix.lower() == ".xlsm",
            }
        finally:
            _close_workbook(wb)

    def tool_excel_list_named_ranges(self, file_path: str) -> dict[str, Any]:
        """List all named ranges (defined names) in a workbook.

        Named ranges are useful for understanding form fields and data references.
        Includes both workbook-scoped and sheet-scoped names.

        Example:
            excel_list_named_ranges(file_path="template.xlsm")

        Args:
            file_path: Path to the .xlsx or .xlsm file

        Returns:
            Dictionary with named ranges, their references, and scopes
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            names = []
            for defined_name in wb.defined_names.values():
                scope = "Workbook"
                if defined_name.localSheetId is not None:
                    try:
                        scope = f"Sheet: {wb.sheetnames[defined_name.localSheetId]}"
                    except IndexError:
                        scope = f"Sheet ID: {defined_name.localSheetId}"

                names.append({
                    "name": defined_name.name,
                    "refers_to": defined_name.attr_text,
                    "scope": scope,
                    "hidden": getattr(defined_name, "hidden", False),
                })

            return {
                "file": path.name,
                "count": len(names),
                "named_ranges": names,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_list_tables(
        self, file_path: str, sheet_name: str | None = None
    ) -> dict[str, Any]:
        """List all tables (ListObjects) in a workbook or specific sheet.

        Excel tables provide structured data with headers, auto-filtering,
        and support for calculated columns.

        Example:
            excel_list_tables(file_path="workbook.xlsx")
            excel_list_tables(file_path="workbook.xlsx", sheet_name="Data")

        Args:
            file_path: Path to the .xlsx or .xlsm file
            sheet_name: Optional sheet name to filter tables

        Returns:
            Dictionary with table information including ranges and columns
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            tables = []
            sheets_to_check = [sheet_name] if sheet_name else wb.sheetnames

            for sname in sheets_to_check:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]
                for table_name in ws.tables:
                    table = ws.tables[table_name]
                    columns = [col.name for col in table.tableColumns]
                    tables.append({
                        "name": table_name,
                        "display_name": table.displayName,
                        "sheet": sname,
                        "range": table.ref,
                        "columns": columns,
                        "has_totals_row": table.totalsRowShown,
                        "style": table.tableStyleInfo.name if table.tableStyleInfo else None,
                    })

            return {
                "file": path.name,
                "table_count": len(tables),
                "tables": tables,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_list_merged_cells(
        self,
        file_path: str,
        sheet_name: str | None = None,
    ) -> dict[str, Any]:
        """List all merged cell regions in a workbook or specific sheet.

        Merged cells are important to understand before patching, as only
        the top-left cell of a merged region should be written to.

        Example:
            excel_list_merged_cells(file_path="template.xlsx")
            excel_list_merged_cells(file_path="template.xlsx", sheet_name="Summary")

        Args:
            file_path: Path to the .xlsx or .xlsm file
            sheet_name: Optional sheet name to filter (default: all sheets)

        Returns:
            Dictionary with merged cell ranges per sheet
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            sheets_to_check = [sheet_name] if sheet_name else wb.sheetnames
            result = {}
            total_count = 0

            for sname in sheets_to_check:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]
                merged_ranges = []
                for merged_range in ws.merged_cells.ranges:
                    coord = merged_range.coord
                    top_left = coord.split(":")[0] if ":" in coord else coord
                    merged_ranges.append({
                        "range": str(merged_range),
                        "top_left": top_left,
                        "min_row": merged_range.min_row,
                        "max_row": merged_range.max_row,
                        "min_col": merged_range.min_col,
                        "max_col": merged_range.max_col,
                    })
                if merged_ranges:
                    result[sname] = merged_ranges
                    total_count += len(merged_ranges)

            return {
                "file": path.name,
                "total_merged_regions": total_count,
                "by_sheet": result,
            }
        finally:
            _close_workbook(wb)

    # =========================================================================
    # CHART OPERATIONS
    # =========================================================================

    def tool_excel_add_chart(
        self,
        file_path: str,
        data_range: str,
        chart_type: str = "line",
        sheet_name: str | None = None,
        title: str | None = None,
        position: str = "E2",
        has_header: bool = True,
        use_first_column_as_categories: bool = True,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """Add a chart to an Excel worksheet.

        Example:
            excel_add_chart(
                file_path="report.xlsx",
                data_range="Sheet1!A1:D8",
                chart_type="line",
                title="Growth Forecast",
                position="F2"
            )

        Args:
            file_path: Path to the .xlsx or .xlsm file
            data_range: Range containing data (e.g., "A1:D10" or "Sheet1!A1:D10")
            chart_type: "line", "bar", "column", or "pie"
            sheet_name: Optional sheet name (overrides sheet in data_range)
            title: Optional chart title
            position: Top-left anchor cell for the chart (default "E2")
            has_header: Treat first row as header for series names
            use_first_column_as_categories: Use first column as category labels
            output_path: Optional output path (defaults to overwriting input)

        Returns:
            Status dictionary with chart details
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}

        try:
            parsed_sheet, range_ref = _parse_range_reference(data_range)
            target_sheet = sheet_name or parsed_sheet or wb.active.title

            if target_sheet not in wb.sheetnames:
                return {"error": f"Sheet not found: {target_sheet}"}

            ws = wb[target_sheet]
            min_col, min_row, max_col, max_row = range_boundaries(range_ref)

            if use_first_column_as_categories and min_col == max_col:
                return {"error": "Data range must include at least two columns when using categories."}

            chart_type_norm = chart_type.lower().strip()
            if chart_type_norm == "line":
                chart = LineChart()
            elif chart_type_norm in ("bar", "column"):
                chart = BarChart()
                if chart_type_norm == "column":
                    chart.type = "col"
            elif chart_type_norm == "pie":
                chart = PieChart()
            else:
                return {"error": "Unsupported chart_type. Use line, bar, column, or pie."}

            data_min_col = min_col + 1 if use_first_column_as_categories else min_col
            data_ref = Reference(
                ws,
                min_col=data_min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
            chart.add_data(data_ref, titles_from_data=has_header)

            if use_first_column_as_categories:
                cat_min_row = min_row + 1 if has_header else min_row
                categories = Reference(
                    ws,
                    min_col=min_col,
                    min_row=cat_min_row,
                    max_row=max_row,
                )
                chart.set_categories(categories)

            if title:
                chart.title = title

            ws.add_chart(chart, position)

            save_path = output_path or file_path
            wb.save(save_path)
            return {
                "success": True,
                "file": save_path,
                "sheet": target_sheet,
                "chart_type": chart_type_norm,
                "position": position,
                "data_range": range_ref,
            }
        finally:
            _close_workbook(wb)

    # =========================================================================
    # COMMENT OPERATIONS
    # =========================================================================

    def tool_excel_add_comment(
        self,
        file_path: str,
        cell_ref: str,
        text: str,
        author: str | None = None,
        sheet_name: str | None = None,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """Add a comment to a cell.

        If the cell already has a comment, the new text is appended.
        Comments are useful for review notes, explanations, or audit trails.

        Example:
            excel_add_comment(
                file_path="workbook.xlsx",
                cell_ref="B5",
                text="Verify this value with finance team",
                author="Reviewer"
            )

        Args:
            file_path: Path to the .xlsx or .xlsm file
            cell_ref: Cell reference (e.g., 'B5', 'Sheet1!C10')
            text: Comment text to add
            author: Author name for the comment (default from environment)
            sheet_name: Optional sheet name (overrides sheet in cell_ref)
            output_path: Optional output path (defaults to overwriting input)

        Returns:
            Status dictionary with comment details
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        author = author or getattr(self, "_comment_author", DEFAULT_AUTHOR)

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            # Parse reference for sheet name
            parsed_sheet, cell_address = _parse_cell_reference(cell_ref)
            target_sheet = sheet_name or parsed_sheet or wb.active.title

            if target_sheet not in wb.sheetnames:
                return {"error": f"Sheet not found: {target_sheet}"}

            ws = wb[target_sheet]
            cell = ws[cell_address]

            had_existing = cell.comment is not None
            if cell.comment:
                cell.comment.text += f"\n\n{text}"
            else:
                cell.comment = Comment(text, author)
                cell.comment.width = 250
                cell.comment.height = 100

            # Save
            save_path = output_path or file_path
            try:
                wb.save(save_path)
            except Exception as e:
                return {"error": f"Failed to save workbook: {e}"}

            return {
                "success": True,
                "sheet": target_sheet,
                "cell": cell_address,
                "author": author,
                "appended": had_existing,
                "file": save_path,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_get_comments(
        self,
        file_path: str,
        sheet_name: str | None = None,
    ) -> dict[str, Any]:
        """Get all comments from a workbook or specific sheet.

        Retrieves cell comments with their authors and text content.

        Example:
            excel_get_comments(file_path="workbook.xlsx")
            excel_get_comments(file_path="workbook.xlsx", sheet_name="Data")

        Args:
            file_path: Path to the .xlsx or .xlsm file
            sheet_name: Optional sheet name to filter (default: all sheets)

        Returns:
            Dictionary with comments organized by sheet
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            sheets_to_check = [sheet_name] if sheet_name else wb.sheetnames
            result = {}
            total_count = 0

            for sname in sheets_to_check:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]
                comments = []
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.comment:
                            comments.append({
                                "cell": cell.coordinate,
                                "author": cell.comment.author,
                                "text": cell.comment.text,
                                "cell_value": cell.value,
                            })
                if comments:
                    result[sname] = comments
                    total_count += len(comments)

            return {
                "file": path.name,
                "total_comments": total_count,
                "by_sheet": result,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_delete_comment(
        self,
        file_path: str,
        cell_ref: str,
        sheet_name: str | None = None,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """Delete a comment from a specific cell.

        Args:
            file_path: Path to the .xlsx or .xlsm file
            cell_ref: Cell reference (e.g., 'B5', 'Sheet1!C10')
            sheet_name: Optional sheet name (overrides sheet in cell_ref)
            output_path: Optional output path (defaults to overwriting input)

        Returns:
            Status dictionary with deletion details
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            parsed_sheet, cell_address = _parse_cell_reference(cell_ref)
            target_sheet = sheet_name or parsed_sheet or wb.active.title

            if target_sheet not in wb.sheetnames:
                return {"error": f"Sheet not found: {target_sheet}"}

            ws = wb[target_sheet]
            cell = ws[cell_address]

            if cell.comment is None:
                return {
                    "error": f"No comment found at {target_sheet}!{cell_address}",
                }

            cell.comment = None

            save_path = output_path or file_path
            try:
                wb.save(save_path)
            except Exception as e:
                return {"error": f"Failed to save workbook: {e}"}

            return {
                "success": True,
                "sheet": target_sheet,
                "cell": cell_address,
                "file": save_path,
                "message": "Comment deleted",
            }
        finally:
            _close_workbook(wb)

    # =========================================================================
    # CELL AND RANGE OPERATIONS
    # =========================================================================

    def tool_excel_get_range(
        self,
        file_path: str,
        range_ref: str,
        sheet_name: str | None = None,
        include_formulas: bool = False,
    ) -> dict[str, Any]:
        """Read a range of cells from an Excel workbook.

        Can read single cells or rectangular ranges. Optionally includes formulas.

        Example:
            excel_get_range(file_path="data.xlsx", range_ref="A1:D10")
            excel_get_range(file_path="data.xlsx", range_ref="Sheet2!B5")
            excel_get_range(file_path="data.xlsx", range_ref="B5:B20", include_formulas=True)

        Args:
            file_path: Path to the .xlsx or .xlsm file
            range_ref: Cell or range reference (e.g., 'A1', 'B2:D10', 'Sheet1!A1')
            sheet_name: Optional sheet name (overrides sheet in range_ref)
            include_formulas: Return formulas instead of computed values

        Returns:
            Dictionary with cell values organized by rows
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            # Load with data_only=False to get formulas, or True for values
            wb = _load_workbook_for_path(file_path, data_only=not include_formulas)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            # Parse reference for sheet name
            parsed_sheet, cell_ref = _parse_cell_reference(range_ref)
            target_sheet = sheet_name or parsed_sheet or wb.active.title

            if target_sheet not in wb.sheetnames:
                return {"error": f"Sheet not found: {target_sheet}"}

            ws = wb[target_sheet]

            try:
                min_row, min_col, max_row, max_col = _get_range_bounds(cell_ref)
            except ValueError as e:
                return {"error": str(e)}

            rows = []
            for row_idx in range(min_row, max_row + 1):
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is not None:
                        row_data.append({
                            "ref": cell.coordinate,
                            "value": value,
                            "type": type(value).__name__,
                        })
                    else:
                        row_data.append({
                            "ref": cell.coordinate,
                            "value": None,
                            "type": "empty",
                        })
                rows.append(row_data)

            return {
                "sheet": target_sheet,
                "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
                "row_count": len(rows),
                "col_count": max_col - min_col + 1,
                "data": rows,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_patch_cell(
        self,
        file_path: str,
        cell_ref: str,
        value: Any,
        sheet_name: str | None = None,
        author: str | None = None,
        log_change: bool = True,
        highlight: bool = False,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """Update a single cell value with change tracking.

        Changes are logged to a _ChangeLog sheet and optionally highlighted.
        Preserves formulas, formatting, and VBA code.

        Note on merged cells: When updating a merged cell region, write to the
        top-left cell only. The value will span the entire merged area.

        Example:
            excel_patch_cell(
                file_path="form.xlsx",
                cell_ref="B5",
                value="Contoso Ltd",
                sheet_name="ECIF Work Scope (E)"
            )

        Args:
            file_path: Path to the .xlsx or .xlsm file
            cell_ref: Cell reference (e.g., 'B5', 'Sheet1!C10')
            value: New value to set (strings starting with '=' are treated as formulas)
            sheet_name: Optional sheet name (overrides sheet in cell_ref)
            author: Author name for change log (default from environment)
            log_change: Whether to log the change (default True)
            highlight: Whether to highlight the changed cell (default False to preserve formatting)
            output_path: Optional output path (defaults to overwriting input)

        Returns:
            Status dictionary with old and new values
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        author = author or DEFAULT_AUTHOR


        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            # Parse reference for sheet name
            parsed_sheet, cell_address = _parse_cell_reference(cell_ref)
            target_sheet = sheet_name or parsed_sheet or wb.active.title

            if target_sheet not in wb.sheetnames:
                return {"error": f"Sheet not found: {target_sheet}"}

            ws = wb[target_sheet]
            cell = ws[cell_address]
            old_value = cell.value

            # Set new value with type coercion
            coerced_value = _set_cell_with_coercion(cell, value, auto_height=True)

            # Auto-adjust row height for multi-line content
            _auto_row_height(ws, cell.row, cell=cell)

            # Log and highlight if requested
            if log_change:
                _log_change(wb, target_sheet, cell_address, old_value, coerced_value, author)

            if highlight:
                _highlight_cell(cell)

            # Save
            save_path = output_path or file_path
            try:
                wb.save(save_path)
            except Exception as e:
                return {"error": f"Failed to save workbook: {e}"}

            return {
                "success": True,
                "sheet": target_sheet,
                "cell": cell_address,
                "old_value": old_value,
                "new_value": coerced_value,
                "value_type": type(coerced_value).__name__,
                "logged": log_change,
                "highlighted": highlight,
                "file": save_path,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_patch_range(
        self,
        file_path: str,
        range_ref: str,
        values: list[list[Any]],
        sheet_name: str | None = None,
        author: str | None = None,
        log_changes: bool = True,
        highlight: bool = False,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """Update a range of cells with change tracking.

        Values should be a 2D array matching the range dimensions.
        Changes are logged to a _ChangeLog sheet.

        Note on merged cells: Merged cell regions within the range will have
        their value set on the top-left cell. Avoid including partial merged
        regions in the range.

        Example:
            excel_patch_range(
                file_path="data.xlsx",
                range_ref="B2:D4",
                values=[
                    ["Alice", 100, "Active"],
                    ["Bob", 200, "Inactive"],
                    ["Carol", 150, "Active"]
                ]
            )

        Args:
            file_path: Path to the .xlsx or .xlsm file
            range_ref: Range reference (e.g., 'B2:D4', 'Sheet1!A1:C10')
            values: 2D array of values to set (strings starting with '=' are treated as formulas)
            sheet_name: Optional sheet name (overrides sheet in range_ref)
            author: Author name for change log
            log_changes: Whether to log changes (default True)
            highlight: Whether to highlight changed cells (default False to preserve formatting)
            output_path: Optional output path

        Returns:
            Status dictionary with change count
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        author = author or DEFAULT_AUTHOR


        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}
        try:
            # Parse reference
            parsed_sheet, cell_ref = _parse_cell_reference(range_ref)
            target_sheet = sheet_name or parsed_sheet or wb.active.title

            if target_sheet not in wb.sheetnames:
                return {"error": f"Sheet not found: {target_sheet}"}

            ws = wb[target_sheet]

            try:
                min_row, min_col, max_row, max_col = _get_range_bounds(cell_ref)
            except ValueError as e:
                return {"error": str(e)}

            # Validate dimensions
            expected_rows = max_row - min_row + 1
            expected_cols = max_col - min_col + 1

            if len(values) != expected_rows:
                return {
                    "error": f"Row count mismatch: expected {expected_rows}, got {len(values)}"
                }

            changes = 0
            rows_to_adjust = set()  # Track rows that need height adjustment

            for row_idx, row_values in enumerate(values):
                if len(row_values) != expected_cols:
                    return {
                        "error": f"Column count mismatch in row {row_idx + 1}: "
                        f"expected {expected_cols}, got {len(row_values)}"
                    }

                for col_idx, new_value in enumerate(row_values):
                    cell = ws.cell(row=min_row + row_idx, column=min_col + col_idx)
                    old_value = cell.value

                    # Use coercion for value setting
                    coerced_value = _coerce_value(new_value)

                    if old_value != coerced_value:
                        _set_cell_with_coercion(cell, new_value, auto_height=True)
                        changes += 1
                        rows_to_adjust.add(min_row + row_idx)

                        if log_changes:
                            _log_change(
                                wb, target_sheet, cell.coordinate, old_value, coerced_value, author
                            )

                        if highlight:
                            _highlight_cell(cell)

            # Auto-adjust row heights for all modified rows
            for row_num in rows_to_adjust:
                _auto_row_height(ws, row_num)

            # Save
            save_path = output_path or file_path
            try:
                wb.save(save_path)
            except Exception as e:
                return {"error": f"Failed to save workbook: {e}"}

            return {
                "success": True,
                "sheet": target_sheet,
                "range": cell_ref,
                "cells_changed": changes,
                "rows_adjusted": len(rows_to_adjust),
                "logged": log_changes,
                "file": save_path,
            }
        finally:
            _close_workbook(wb)

    # =========================================================================
    # TABLE OPERATIONS
    # =========================================================================

    def tool_excel_get_table(
        self,
        file_path: str,
        table_name: str,
        include_headers: bool = True,
    ) -> dict[str, Any]:
        """Read data from an Excel table by name.

        Retrieves all rows from a named table (ListObject) including
        optional headers.

        Example:
            excel_get_table(file_path="workbook.xlsx", table_name="Milestones")

        Args:
            file_path: Path to the .xlsx or .xlsm file
            table_name: Name of the table
            include_headers: Include header row in output (default True)

        Returns:
            Dictionary with table data as list of rows
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=True)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}

        try:
            # Find the table
            table_info = None
            table_ws = None
            for sname in wb.sheetnames:
                ws = wb[sname]
                if table_name in ws.tables:
                    table_info = ws.tables[table_name]
                    table_ws = ws
                    break

            if not table_info:
                return {"error": f"Table not found: {table_name}"}

            # Parse table range
            try:
                min_row, min_col, max_row, max_col = _get_range_bounds(table_info.ref)
            except ValueError as e:
                return {"error": str(e)}

            columns = [col.name for col in table_info.tableColumns]
            rows = []

            start_row = min_row if include_headers else min_row + 1
            for row_idx in range(start_row, max_row + 1):
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell = table_ws.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                rows.append(row_data)

            return {
                "table_name": table_name,
                "sheet": table_ws.title,
                "range": table_info.ref,
                "columns": columns,
                "row_count": len(rows),
                "data": rows,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_append_table_row(
        self,
        file_path: str,
        table_name: str,
        row_data: dict[str, Any],
        author: str | None = None,
        output_path: str | None = None,
        mode: Literal["best_effort", "safe", "strict", "dry_run"] = "best_effort",
    ) -> dict[str, Any]:
        """Append a new row to an Excel table.

        Automatically expands the table range and maintains formatting.
        Use column names as keys in row_data.

        Example:
            excel_append_table_row(
                file_path="workbook.xlsx",
                table_name="Milestones",
                row_data={
                    "Milestone #": "5",
                    "Description": "Phase 2 Complete",
                    "Due Date": "2026-06-30"
                }
            )

        Args:
            file_path: Path to the .xlsx or .xlsm file
            table_name: Name of the table
            row_data: Dictionary mapping column names to values
            author: Author name for change log
            output_path: Optional output path

        Returns:
            Status dictionary with new row number
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        author = author or DEFAULT_AUTHOR

        if mode == "safe" and (output_path is None or Path(output_path).resolve() == path.resolve()):
            return {
                "success": False,
                "mode": mode,
                "status": "failed",
                "warnings": ["safe mode requires an explicit output_path different from the source file."],
                "matched_targets": [],
                "unmatched_targets": [{"target": f"table:{table_name}", "reason": "safe_mode_requires_distinct_output_path"}],
                "skipped_targets": [],
                "diagnostics": {"table": table_name},
                "next_tools": ["office_help", "office_table", "office_template"],
            }

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}

        try:
            # Find the table
            table_info = None
            table_ws = None
            for sname in wb.sheetnames:
                ws = wb[sname]
                if table_name in ws.tables:
                    table_info = ws.tables[table_name]
                    table_ws = ws
                    break

            if not table_info:
                return {"error": f"Table not found: {table_name}"}

            # Parse current table range
            try:
                min_row, min_col, max_row, max_col = _get_range_bounds(table_info.ref)
            except ValueError as e:
                return {"error": str(e)}

            # Get column mapping
            columns = {col.name: idx for idx, col in enumerate(table_info.tableColumns)}

            # Calculate new row
            new_row = max_row + 1
            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_row}"
            matched_columns = [col for col in row_data.keys() if col in columns]
            unmatched_columns = [
                {"target": f"column:{col}", "reason": "column_not_found"}
                for col in row_data.keys() if col not in columns
            ]

            if mode == "strict" and unmatched_columns:
                return {
                    "success": False,
                    "mode": mode,
                    "status": "failed",
                    "warnings": ["strict mode requires every provided column to match the Excel table."],
                    "matched_targets": [],
                    "unmatched_targets": unmatched_columns,
                    "skipped_targets": [],
                    "diagnostics": {"table": table_name, "columns_available": list(columns.keys())},
                    "next_tools": ["office_inspect", "office_table", "office_help"],
                }

            if mode == "dry_run":
                diag = build_mutation_diagnostics(
                    matched_targets=[{"target": f"table:{table_name}", "columns_filled": matched_columns, "row": new_row}] if matched_columns else [],
                    unmatched_targets=unmatched_columns,
                    warnings=["dry_run mode does not write output files."] + (["Some provided columns were not found in the Excel table."] if unmatched_columns else []),
                    diagnostics={
                        "table": table_name,
                        "row_index": new_row,
                        "columns_available": list(columns.keys()),
                        "columns_filled": matched_columns,
                        "predicted_new_range": new_ref,
                    },
                    next_tools=["office_read", "office_table", "office_audit", "office_inspect"],
                )
                return {
                    **diag,
                    "mode": mode,
                    "table": table_name,
                    "new_row": new_row,
                    "new_range": new_ref,
                    "columns_filled": list(row_data.keys()),
                    "file": output_path or file_path,
                }

            # Write cell values with type coercion
            for col_name, value in row_data.items():
                if col_name not in columns:
                    continue
                col_idx = min_col + columns[col_name]
                cell = table_ws.cell(row=new_row, column=col_idx)
                coerced_value = _set_cell_with_coercion(cell, value, auto_height=True)

                # Log the change
                _log_change(wb, table_ws.title, cell.coordinate, None, coerced_value, author)

            # Auto-adjust row height for the new row
            _auto_row_height(table_ws, new_row)

            # Expand table range
            table_info.ref = new_ref

            # Save
            save_path = output_path or file_path
            try:
                wb.save(save_path)
            except Exception as e:
                return {"error": f"Failed to save workbook: {e}"}
            diag = build_mutation_diagnostics(
                matched_targets=[{"target": f"table:{table_name}", "columns_filled": matched_columns, "row": new_row}] if matched_columns else [],
                unmatched_targets=unmatched_columns,
                warnings=["Some provided columns were not found in the Excel table."] if unmatched_columns else [],
                diagnostics={
                    "table": table_name,
                    "row_index": new_row,
                    "columns_available": list(columns.keys()),
                    "columns_filled": matched_columns,
                },
                next_tools=["office_read", "office_table", "office_audit", "office_inspect"],
            )
            return {
                **diag,
                "mode": mode,
                "table": table_name,
                "new_row": new_row,
                "new_range": new_ref,
                "columns_filled": list(row_data.keys()),
                "file": save_path,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_update_table_row(
        self,
        file_path: str,
        table_name: str,
        row_index: int,
        row_data: dict[str, Any],
        author: str | None = None,
        output_path: str | None = None,
        mode: Literal["best_effort", "safe", "strict", "dry_run"] = "best_effort",
    ) -> dict[str, Any]:
        """Update values in an existing table row.

        Row index is 1-based (1 = first data row after headers).

        Example:
            excel_update_table_row(
                file_path="workbook.xlsx",
                table_name="Milestones",
                row_index=2,
                row_data={"Status": "Complete", "Actual Date": "2026-03-15"}
            )

        Args:
            file_path: Path to the .xlsx or .xlsm file
            table_name: Name of the table
            row_index: 1-based row index (1 = first data row)
            row_data: Dictionary mapping column names to new values
            author: Author name for change log
            output_path: Optional output path

        Returns:
            Status dictionary with updated values
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        author = author or DEFAULT_AUTHOR

        if mode == "safe" and (output_path is None or Path(output_path).resolve() == path.resolve()):
            return {
                "success": False,
                "mode": mode,
                "status": "failed",
                "warnings": ["safe mode requires an explicit output_path different from the source file."],
                "matched_targets": [],
                "unmatched_targets": [{"target": f"table:{table_name}/row:{row_index}", "reason": "safe_mode_requires_distinct_output_path"}],
                "skipped_targets": [],
                "diagnostics": {"table": table_name, "row_index": row_index},
                "next_tools": ["office_help", "office_table", "office_template"],
            }

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}

        try:
            # Find the table
            table_info = None
            table_ws = None
            for sname in wb.sheetnames:
                ws = wb[sname]
                if table_name in ws.tables:
                    table_info = ws.tables[table_name]
                    table_ws = ws
                    break

            if not table_info:
                return {"error": f"Table not found: {table_name}"}

            # Parse table range
            try:
                min_row, min_col, max_row, max_col = _get_range_bounds(table_info.ref)
            except ValueError as e:
                return {"error": str(e)}

            # Calculate actual row (header + row_index)
            target_row = min_row + row_index  # min_row is header, so +1 for data

            if target_row > max_row:
                return {"error": f"Row index {row_index} out of range (max: {max_row - min_row})"}

            # Get column mapping
            columns = {col.name: idx for idx, col in enumerate(table_info.tableColumns)}
            unmatched_columns = [
                {"target": f"column:{col}", "reason": "column_not_found"}
                for col in row_data.keys() if col not in columns
            ]

            if mode == "strict" and unmatched_columns:
                return {
                    "success": False,
                    "mode": mode,
                    "status": "failed",
                    "warnings": ["strict mode requires every provided column to match the Excel table."],
                    "matched_targets": [],
                    "unmatched_targets": unmatched_columns,
                    "skipped_targets": [],
                    "diagnostics": {"table": table_name, "row_index": row_index, "columns_available": list(columns.keys())},
                    "next_tools": ["office_inspect", "office_table", "office_help"],
                }

            # Update cells with type coercion
            updates = []
            for col_name, new_value in row_data.items():
                if col_name not in columns:
                    continue
                col_idx = min_col + columns[col_name]
                cell = table_ws.cell(row=target_row, column=col_idx)
                old_value = cell.value
                coerced_value = _set_cell_with_coercion(cell, new_value, auto_height=True)

                updates.append({
                    "column": col_name,
                    "old_value": old_value,
                    "new_value": coerced_value,
                    "value_type": type(coerced_value).__name__,
                })

                if mode != "dry_run":
                    _log_change(wb, table_ws.title, cell.coordinate, old_value, coerced_value, author)
                    _highlight_cell(cell)

            if mode == "dry_run":
                diag = build_mutation_diagnostics(
                    matched_targets=[{"target": f"table:{table_name}/row:{row_index}", "updates": updates}] if updates else [],
                    unmatched_targets=unmatched_columns,
                    warnings=["dry_run mode does not write output files."] + (["Some provided columns were not found in the Excel table."] if unmatched_columns else []),
                    diagnostics={
                        "table": table_name,
                        "row_index": row_index,
                        "columns_available": list(columns.keys()),
                        "updates": updates,
                    },
                    next_tools=["office_read", "office_table", "office_audit", "office_inspect"],
                )
                return {
                    **diag,
                    "mode": mode,
                    "table": table_name,
                    "row_index": row_index,
                    "updates": updates,
                    "file": output_path or file_path,
                }

            # Auto-adjust row height for the updated row
            _auto_row_height(table_ws, target_row)

            # Save
            save_path = output_path or file_path
            try:
                wb.save(save_path)
            except Exception as e:
                return {"error": f"Failed to save workbook: {e}"}
            diag = build_mutation_diagnostics(
                matched_targets=[{"target": f"table:{table_name}/row:{row_index}", "updates": updates}] if updates else [],
                unmatched_targets=unmatched_columns,
                warnings=["Some provided columns were not found in the Excel table."] if unmatched_columns else [],
                diagnostics={
                    "table": table_name,
                    "row_index": row_index,
                    "columns_available": list(columns.keys()),
                    "updates": updates,
                },
                next_tools=["office_read", "office_table", "office_audit", "office_inspect"],
            )
            return {
                **diag,
                "mode": mode,
                "table": table_name,
                "row_index": row_index,
                "updates": updates,
                "file": save_path,
            }
        finally:
            _close_workbook(wb)

    # =========================================================================
    # PLACEHOLDER OPERATIONS
    # =========================================================================

    def tool_excel_replace_placeholders(
        self,
        file_path: str,
        replacements: dict[str, Any],
        sheet_names: list[str] | None = None,
        author: str | None = None,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """Replace placeholder patterns throughout the workbook.

        Searches for placeholders like <Customer Name> or [TBD] and replaces
        them with provided values. All changes are logged.

        Example:
            excel_replace_placeholders(
                file_path="template.xlsx",
                replacements={
                    "<Customer Name>": "Contoso Ltd",
                    "<Project Name>": "Cloud Migration",
                    "[TBD]": "Q1 2026"
                }
            )

        Args:
            file_path: Path to the .xlsx or .xlsm file
            replacements: Dictionary mapping placeholders to replacement values
            sheet_names: Optional list of sheets to process (default: all visible)
            author: Author name for change log
            output_path: Optional output path

        Returns:
            Status dictionary with replacement counts per placeholder
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        author = author or DEFAULT_AUTHOR

        try:
            wb = _load_workbook_for_path(file_path, data_only=False)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}

        try:
            sheets_to_process = sheet_names or [
                s for s in wb.sheetnames if wb[s].sheet_state == "visible"
            ]

            counts = dict.fromkeys(replacements, 0)
            total = 0

            for sname in sheets_to_process:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None or not isinstance(cell.value, str):
                            continue

                        original = cell.value
                        modified = original

                        for placeholder, replacement in replacements.items():
                            if placeholder in modified:
                                modified = modified.replace(placeholder, str(replacement))
                                counts[placeholder] += 1
                                total += 1

                        if modified != original:
                            cell.value = modified
                            _log_change(wb, sname, cell.coordinate, original, modified, author)
                            _highlight_cell(cell)

            # Save
            save_path = output_path or file_path
            try:
                wb.save(save_path)
            except Exception as e:
                return {"error": f"Failed to save workbook: {e}"}

            return {
                "success": True,
                "total_replacements": total,
                "by_placeholder": counts,
                "sheets_processed": sheets_to_process,
                "file": save_path,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_audit_placeholders(
        self,
        file_path: str,
        patterns: list[str] | None = None,
        sheet_names: list[str] | None = None,
    ) -> dict[str, Any]:
        """Audit workbook for unfilled placeholders.

        Searches for patterns like <...>, [...], [TBD], etc. to identify
        cells that still need to be filled.

        Example:
            excel_audit_placeholders(file_path="form.xlsx")

            excel_audit_placeholders(
                file_path="form.xlsx",
                patterns=["<Customer Name>", "[TBD]", "PLACEHOLDER"]
            )

        Args:
            file_path: Path to the .xlsx or .xlsm file
            patterns: Optional list of specific patterns to search for
            sheet_names: Optional list of sheets to check

        Returns:
            Audit report with found placeholders and their locations
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=True)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}

        try:
            # Default placeholder patterns
            if patterns is None:
                patterns = [
                    r"<[^>]+>",  # <anything>
                    r"\[[^\]]*TBD[^\]]*\]",  # [TBD], [To Be Determined], etc.
                    r"\[insert[^\]]*\]",  # [insert ...]
                    r"\[enter[^\]]*\]",  # [enter ...]
                    r"\[select[^\]]*\]",  # [select ...]
                ]
                use_regex = True
            else:
                use_regex = False

            sheets_to_check = sheet_names or [
                s for s in wb.sheetnames if wb[s].sheet_state == "visible"
            ]

            findings = []

            for sname in sheets_to_check:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None or not isinstance(cell.value, str):
                            continue

                        text = str(cell.value)

                        for pattern in patterns:
                            if use_regex:
                                matches = re.findall(pattern, text, re.IGNORECASE)
                                for match in matches:
                                    findings.append({
                                        "sheet": sname,
                                        "cell": cell.coordinate,
                                        "pattern": pattern,
                                        "match": match,
                                        "context": text[:100],
                                    })
                            elif pattern.lower() in text.lower():
                                findings.append({
                                    "sheet": sname,
                                    "cell": cell.coordinate,
                                    "pattern": pattern,
                                    "match": pattern,
                                    "context": text[:100],
                                })

            # Group by pattern
            by_pattern: dict[str, list[dict]] = {}
            for finding in findings:
                p = finding["pattern"]
                if p not in by_pattern:
                    by_pattern[p] = []
                by_pattern[p].append(finding)

            return {
                "file": path.name,
                "total_found": len(findings),
                "by_pattern": {p: len(f) for p, f in by_pattern.items()},
                "sheets_checked": sheets_to_check,
                "findings": findings,
                "status": "clean" if not findings else "needs_attention",
            }
        finally:
            _close_workbook(wb)

    # =========================================================================
    # COPY AND TEMPLATE OPERATIONS
    # =========================================================================

    def tool_excel_copy_template(
        self,
        template_path: str,
        output_path: str,
    ) -> dict[str, Any]:
        """Copy an Excel template to start a new document.

        Preserves all content including VBA macros, formatting, and data validation.

        Example:
            excel_copy_template(
                template_path="templates/ECIF Request Work Scope.xlsm",
                output_path="04. Artifacts/contoso-ecif.xlsm"
            )

        Args:
            template_path: Path to the template file
            output_path: Destination path for the new file

        Returns:
            Status with source and destination paths
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        src_path = Path(template_path)
        if not src_path.exists():
            return {"error": f"Template not found: {template_path}"}

        try:
            # Load with VBA preservation
            wb = _load_workbook_for_path(template_path, data_only=False)
            try:
                wb.save(output_path)
            finally:
                _close_workbook(wb)
        except Exception as e:
            return {"error": f"Failed to copy template: {e}"}

        return {
            "success": True,
            "source": template_path,
            "destination": output_path,
            "message": f"Template copied successfully to {output_path}",
        }

    def tool_excel_get_change_log(self, file_path: str) -> dict[str, Any]:
        """Retrieve the change log from a workbook.

        Returns all logged changes from the _ChangeLog sheet if it exists.

        Example:
            excel_get_change_log(file_path="edited.xlsx")

        Args:
            file_path: Path to the .xlsx or .xlsm file

        Returns:
            Dictionary with change log entries
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path, data_only=True)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}

        try:
            if "_ChangeLog" not in wb.sheetnames:
                return {
                    "file": path.name,
                    "has_change_log": False,
                    "entries": [],
                    "message": "No change log found in this workbook",
                }

            ws = wb["_ChangeLog"]
            entries = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                entries.append({
                    "timestamp": row[0],
                    "sheet": row[1],
                    "cell": row[2],
                    "old_value": row[3],
                    "new_value": row[4],
                    "author": row[5] if len(row) > 5 else None,
                })

            return {
                "file": path.name,
                "has_change_log": True,
                "entry_count": len(entries),
                "entries": entries,
            }
        finally:
            _close_workbook(wb)

    def tool_excel_add_sheet(
        self,
        file_path: str,
        sheet_name: str,
        position: str | None = None,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """Add a new sheet to an Excel workbook.

        Creates a new empty sheet in the workbook at the specified position.
        The sheet can be inserted at the start, end, or after a specific existing sheet.

        Example:
            excel_add_sheet(file_path="data.xlsx", sheet_name="Summary")
            excel_add_sheet(file_path="data.xlsx", sheet_name="NewSheet", position="start")
            excel_add_sheet(file_path="data.xlsx", sheet_name="Details", position="Sheet1")

        Args:
            file_path: Path to the .xlsx or .xlsm file
            sheet_name: Name for the new sheet
            position: Where to insert - 'start', 'end' (default), or name of sheet to insert after
            output_path: Optional output path (defaults to overwriting input file)

        Returns:
            Dictionary with success status and sheet details
        """
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        if not sheet_name or not sheet_name.strip():
            return {"error": "Sheet name is required"}

        sheet_name = sheet_name.strip()

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            wb = _load_workbook_for_path(file_path)
        except Exception as e:
            return {"error": f"Failed to load workbook: {e}"}

        # Check for duplicate sheet name
        if sheet_name in wb.sheetnames:
            return {"error": f"Sheet '{sheet_name}' already exists in the workbook"}

        # Determine position
        pos = position.strip().lower() if position else "end"
        insert_index = None

        if pos == "end" or not position:
            insert_index = len(wb.sheetnames)
        elif pos == "start":
            insert_index = 0
        else:
            # Look for sheet name to insert after
            for i, sname in enumerate(wb.sheetnames):
                if sname.lower() == pos:
                    insert_index = i + 1
                    break
            if insert_index is None:
                return {"error": f"Sheet '{position}' not found to insert after"}

        # Create the new sheet
        wb.create_sheet(title=sheet_name, index=insert_index)

        # Determine save path
        save_path = output_path if output_path else file_path
        save_dir = Path(save_path).parent
        if not save_dir.exists():
            save_dir.mkdir(parents=True, exist_ok=True)

        try:
            wb.save(save_path)
        except Exception as e:
            return {"error": f"Failed to save workbook: {e}"}

        # Get final sheet index
        final_index = wb.sheetnames.index(sheet_name)

        return {
            "success": True,
            "file": Path(save_path).name,
            "sheet_name": sheet_name,
            "sheet_index": final_index,
            "total_sheets": len(wb.sheetnames),
            "message": f"Sheet '{sheet_name}' added at position {final_index}",
        }

